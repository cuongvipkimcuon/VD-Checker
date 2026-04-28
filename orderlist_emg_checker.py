import json
import re
import sqlite3
import tkinter as tk
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
import unicodedata

import pandas as pd

try:
    from tkcalendar import DateEntry
except ImportError:
    DateEntry = None  # type: ignore[misc, assignment]


DB_FILE = "orderlist_emg_checker.db"
CONFIG_FILE = "orderlist_emg_checker_config.json"
CUSTOMER_FILTER_EMG = "EMG"
CUSTOMER_FILTER_NON_EMG = "Ngoài EMG"
CUSTOMER_FILTER_ALL = "Tất cả"

# ORDER LIST sheet1: cột Excel H = Production No. (0-based 7), M = Ship date (0-based 12)
OL_COL_PRODUCTION_NO = 7
OL_COL_SHIP_DATE = 12

USER_CONFIRM_OK = "Đã xác nhận đúng"
USER_CONFIRM_BAD = "Đã xác nhận sai"
RULE_CHECK_FIELD_PREFIX = "Quy tắc |"
CASE_CONCLUSION_RULE_BAD = "Sai quy tắc"

CHECK_STATUS_FILTER_VALUES = (
    "Tất cả",
    "Đúng",
    "Lệch",
    CASE_CONCLUSION_RULE_BAD,
    USER_CONFIRM_OK,
    USER_CONFIRM_BAD,
)

# Tiêu chí có thể thiết lập trong Quy tắc (khớp field_name khi so với OL)
RULE_CRITERIA_FIELDS: tuple[str, ...] = (
    "Đơn hàng",
    "Số lượng đơn hàng",
    "Logo",
    "Loại logo",
    "Mã sản phẩm",
    "Tên sản phẩm",
    "Màu sắc",
    "Số thùng (carton)",
    "Size thùng",
    "Số thùng (pallet)",
    "Size pallet",
)


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_key(value: object) -> str:
    base = clean_text(value).lower()
    base = unicodedata.normalize("NFKD", base)
    base = "".join(ch for ch in base if not unicodedata.combining(ch))
    base = base.replace("đ", "d")
    return re.sub(r"[^a-z0-9]", "", base)


def to_number(value: object) -> float | None:
    if pd.isna(value):
        return None
    text = clean_text(value).replace(",", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def mode_value(series: pd.Series) -> str:
    values = [clean_text(v) for v in series if clean_text(v)]
    if not values:
        return ""
    counter = Counter(values)
    return counter.most_common(1)[0][0]


def rule_size_xyz_match(ol_xyz: object, expected: object) -> bool:
    """So chuỗi size dạng a.b.c (OL vs quy tắc); phần OL thiếu so với số phần quy tắc → không khớp."""
    ol_s = clean_text(ol_xyz).replace(",", ".")
    ex_s = clean_text(expected).replace(",", ".")
    if not ex_s:
        return True
    ol_parts = [clean_text(p) for p in ol_s.split(".") if clean_text(p)]
    ex_parts = [clean_text(p) for p in ex_s.split(".") if clean_text(p)]
    if not ex_parts:
        return True
    if len(ol_parts) < len(ex_parts):
        return False
    for o, e in zip(ol_parts[: len(ex_parts)], ex_parts):
        if not almost_equal(o, e) and clean_key(o) != clean_key(e):
            return False
    return True


def rule_compare_ol_expected(field_name: str, ol_raw: object, expected: str, ol_missing: bool) -> tuple[str, str]:
    """So quy tắc: chỉ giá trị OL vs expected. Trả (status_core, adjust_reason)."""
    exp = clean_text(expected)
    if not exp:
        return ("Đúng", "")
    if ol_missing:
        return ("Lệch", "OL không có dữ liệu (quy tắc).")
    ol_disp = clean_text(ol_raw)
    if field_name in ("Size thùng", "Size pallet"):
        ok = rule_size_xyz_match(ol_raw, expected)
        return ("Đúng" if ok else "Lệch", "" if ok else f"OL «{ol_disp}» khác quy tắc «{exp}».")
    if field_name == "Màu sắc":
        ok = normalize_color_name(ol_raw) == normalize_color_name(exp)
        return ("Đúng" if ok else "Lệch", "" if ok else f"OL «{ol_disp}» khác quy tắc «{exp}».")
    if field_name == "Số lượng đơn hàng" or field_name.startswith("Số thùng"):
        st = qty_status_strict(ol_disp, exp)
        return (st, "" if st == "Đúng" else f"OL «{ol_disp}» khác quy tắc «{exp}».")
    ok = almost_equal(ol_raw, exp)
    return ("Đúng" if ok else "Lệch", "" if ok else f"OL «{ol_disp}» khác quy tắc «{exp}».")


def almost_equal(left: object, right: object, tol: float = 1e-6) -> bool:
    l_num = to_number(left)
    r_num = to_number(right)
    if l_num is not None and r_num is not None:
        return abs(l_num - r_num) <= tol
    return clean_key(left) == clean_key(right)


def normalize_logo(value: object) -> str:
    key = clean_key(value)
    if key in {"nologo", "nologo.", "nologo-"}:
        return "NL"
    return clean_text(value)


def logo_type_from_order_logo_text(value: object) -> str:
    """Map logo type text (cột R) thành mã chuẩn 710/720/730 hoặc NO LOGO."""
    key = clean_key(value)
    if not key:
        return "NO LOGO"
    codes: list[str] = []
    if "transfer" in key:
        codes.append("710")
    if "print" in key:
        codes.append("720")
    if "embroi" in key:
        codes.append("730")
    if not codes:
        return "NO LOGO"
    ordered: list[str] = []
    for code in ("710", "720", "730"):
        if code in codes and code not in ordered:
            ordered.append(code)
    return "+".join(ordered)


def logo_type_from_bang_ke_npl(group: pd.DataFrame) -> str:
    """Dựa vào Mã NPL có prefix 710/720/730 (thường ở cột 9)."""
    if group.empty or group.shape[1] <= 9:
        return "NO LOGO"
    codes: list[str] = []
    for v in group.iloc[:, 9]:
        key = clean_key(v)
        if key.startswith("710"):
            codes.append("710")
        elif key.startswith("720"):
            codes.append("720")
        elif key.startswith("730"):
            codes.append("730")
    if not codes:
        return "NO LOGO"
    ordered: list[str] = []
    for code in ("710", "720", "730"):
        if code in codes and code not in ordered:
            ordered.append(code)
    return "+".join(ordered)


def logo_type_to_label(value: object) -> str:
    """Hiển thị thân thiện: 710/720/730 kèm tên loại logo."""
    raw = clean_text(value).upper()
    if not raw or raw == "NO LOGO":
        return "NO LOGO"
    labels = {
        "710": "710 (Transfer)",
        "720": "720 (Print)",
        "730": "730 (Embroidery)",
    }
    parts = [p.strip() for p in raw.split("+") if p.strip()]
    out: list[str] = []
    for p in parts:
        out.append(labels.get(p, p))
    if not out:
        return "NO LOGO"
    return " + ".join(out)


def qty_status(order_qty: object, bang_ke_qty: object) -> str:
    order_num = to_number(order_qty)
    bk_num = to_number(bang_ke_qty)
    if order_num is None or bk_num is None:
        return "Đúng" if almost_equal(order_qty, bang_ke_qty) else "Lệch"
    if abs(order_num - bk_num) <= 1e-6:
        return "Đúng"
    # Rule nghiệp vụ: Bảng kê lớn hơn Order List đúng 1 vẫn chấp nhận.
    if 0 < (bk_num - order_num) <= 1:
        return "Đúng"
    return "Lệch"


def qty_status_strict(order_qty: object, bang_ke_qty: object) -> str:
    """So số lượng nghiêm ngặt: không cho lệch +1."""
    return "Đúng" if almost_equal(order_qty, bang_ke_qty) else "Lệch"


def parse_npl_950_code(code: object) -> tuple[str, str, str]:
    text = clean_text(code)
    if not text:
        return ("", "", "")
    parts = [p.strip() for p in text.split(".")]
    if len(parts) < 4:
        return ("", "", "")
    return (parts[1], parts[2], parts[3])


def parse_npl_color(code: object) -> str:
    text = clean_text(code)
    if not text:
        return ""
    parts = [p.strip() for p in text.split(".")]
    if len(parts) < 2:
        return ""
    return parts[1]


DEFAULT_COLOR_CODE_PAIRS: tuple[tuple[str, str], ...] = (
    ("100", "Black"),
    ("507", "Cam"),
    ("204", "Marine Blue"),
    ("503", "Red"),
    ("800", "White"),
    ("209", "Blue"),
    ("305", "Green"),
    ("702", "Yellow"),
)
COLOR_MAP_CONFIG_KEY = "color_code_pairs"


def color_name_from_code(color_code: str, color_pairs: list[dict[str, str]] | None = None) -> str:
    key = clean_text(color_code)
    if not key:
        return ""
    pairs = color_pairs if color_pairs is not None else load_color_pairs_from_config({})
    names: list[str] = []
    seen: set[str] = set()
    for item in pairs:
        code = clean_text(item.get("code", ""))
        name = clean_text(item.get("name", ""))
        if code != key or not name:
            continue
        name_key = clean_key(name)
        if name_key in seen:
            continue
        seen.add(name_key)
        names.append(name)
    if names:
        return " / ".join(names)
    return key


def normalize_color_name(value: object) -> str:
    key = clean_key(value)
    alias = {
        "black": "black",
        "cam": "cam",
        "orange": "cam",
        "marineblue": "marineblue",
        "red": "red",
        "white": "white",
        "blue": "blue",
        "green": "green",
        "yellow": "yellow",
    }
    return alias.get(key, key)


def load_color_pairs_from_config(config: dict) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    pairs = config.get(COLOR_MAP_CONFIG_KEY, [])
    if isinstance(pairs, list):
        for item in pairs:
            if not isinstance(item, dict):
                continue
            code = clean_text(item.get("code", ""))
            name = clean_text(item.get("name", ""))
            if code and name:
                rows.append({"code": code, "name": name})
    if rows:
        return rows
    for code, name in DEFAULT_COLOR_CODE_PAIRS:
        rows.append({"code": code, "name": name})
    return rows


def build_color_alias_lookup(color_pairs: list[dict[str, str]]) -> dict[str, set[str]]:
    lookup: dict[str, set[str]] = {}
    for item in color_pairs:
        code = clean_text(item.get("code", ""))
        name = clean_text(item.get("name", ""))
        if not code or not name:
            continue
        if code not in lookup:
            lookup[code] = set()
        lookup[code].add(clean_key(name))
        lookup[code].add(normalize_color_name(name))
    return lookup


def color_value_matches_code(value: object, color_code: str, color_alias_lookup: dict[str, set[str]]) -> bool:
    code = clean_text(color_code)
    if not code:
        return False
    aliases = color_alias_lookup.get(code, set())
    if not aliases:
        return False
    return normalize_color_name(value) in aliases or clean_key(value) in aliases


def dedupe_color_pairs(color_pairs: list[dict[str, str]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for item in color_pairs:
        if not isinstance(item, dict):
            continue
        code = clean_text(item.get("code", ""))
        name = clean_text(item.get("name", ""))
        if not code or not name:
            continue
        key = (code, clean_key(name))
        if key in seen:
            continue
        seen.add(key)
        out.append({"code": code, "name": name})
    return out


def load_config() -> dict:
    path = Path(CONFIG_FILE)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_config(data: dict) -> None:
    Path(CONFIG_FILE).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def init_db() -> None:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_type TEXT NOT NULL,
            target_dg TEXT,
            order_file TEXT NOT NULL,
            bang_ke_file TEXT NOT NULL,
            run_label TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS run_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL,
            dg_case_no TEXT NOT NULL,
            field_name TEXT NOT NULL,
            order_value TEXT,
            shipped_value TEXT,
            bang_ke_value TEXT,
            status TEXT NOT NULL,
            FOREIGN KEY (run_id) REFERENCES runs(id) ON DELETE CASCADE
        )
        """
    )
    existing_cols = {row[1] for row in cur.execute("PRAGMA table_info(run_items)").fetchall()}
    if "auto_status" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN auto_status TEXT")
    if "is_adjusted" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN is_adjusted INTEGER DEFAULT 0")
    if "adjusted_at" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN adjusted_at TEXT")
    if "adjust_reason" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN adjust_reason TEXT")
    if "shipped_value" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN shipped_value TEXT")
    if "production_no" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN production_no TEXT DEFAULT ''")
    if "ship_date_display" not in existing_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN ship_date_display TEXT DEFAULT ''")
    existing_runs_cols = {row[1] for row in cur.execute("PRAGMA table_info(runs)").fetchall()}
    if "run_label" not in existing_runs_cols:
        cur.execute("ALTER TABLE runs ADD COLUMN run_label TEXT DEFAULT ''")
    if "trace_orderlist_filename" not in existing_runs_cols:
        cur.execute("ALTER TABLE runs ADD COLUMN trace_orderlist_filename TEXT DEFAULT ''")
    if "trace_shipped_file_mtime" not in existing_runs_cols:
        cur.execute("ALTER TABLE runs ADD COLUMN trace_shipped_file_mtime TEXT DEFAULT ''")
    if "trace_bang_ke_a6" not in existing_runs_cols:
        cur.execute("ALTER TABLE runs ADD COLUMN trace_bang_ke_a6 TEXT DEFAULT ''")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS run_case_meta (
            run_id INTEGER NOT NULL,
            dg_case_no TEXT NOT NULL,
            user_conclusion TEXT DEFAULT '',
            case_note TEXT DEFAULT '',
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (run_id, dg_case_no),
            FOREIGN KEY (run_id) REFERENCES runs(id) ON DELETE CASCADE
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS dg_rule_sets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dg_case_no TEXT NOT NULL UNIQUE,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS dg_rule_criteria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_set_id INTEGER NOT NULL,
            field_name TEXT NOT NULL,
            expected_value TEXT DEFAULT '',
            UNIQUE (rule_set_id, field_name),
            FOREIGN KEY (rule_set_id) REFERENCES dg_rule_sets(id) ON DELETE CASCADE
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pinned_runs (
            run_id INTEGER PRIMARY KEY,
            pinned_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (run_id) REFERENCES runs(id) ON DELETE CASCADE
        )
        """
    )
    conn.commit()
    conn.close()


def trace_orderlist_filename(order_path: str) -> str:
    return clean_text(Path(order_path).name)


def trace_shipped_file_mtime_local(shipped_path: str) -> str:
    """Thời điểm chỉnh sửa file SHIPPED LIST trên đĩa (truy vết)."""
    p = Path(shipped_path)
    if not p.is_file():
        return ""
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    except OSError:
        return ""


def trace_bang_ke_sheet_a6(bk_path: str) -> str:
    """Ô A6 sheet đầu của Bảng Kê (có thể merge) — lấy giá trị hiển thị."""
    p = Path(bk_path)
    if not p.is_file():
        return ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(p), read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            v = ws["A6"].value
            return clean_text(v)
        finally:
            wb.close()
    except Exception:
        try:
            prev = pd.read_excel(str(p), sheet_name=0, header=None, nrows=6, engine="openpyxl")
            if prev.shape[0] >= 6 and prev.shape[1] >= 1:
                return clean_text(prev.iloc[5, 0])
        except Exception:
            pass
        return ""


def find_bang_ke_header_row(file_path: str) -> int:
    preview = pd.read_excel(file_path, sheet_name=0, header=None, nrows=40)
    for idx, row in preview.iterrows():
        row_values = [clean_key(v) for v in row.tolist()]
        line = "|".join(v for v in row_values if v)
        if "soso" in line and "masanpham" in line and "ghichu" in line:
            return int(idx)
    raise ValueError("Không tìm thấy dòng header trong file Bảng Kê.")


def filter_bang_ke_rows_for_dg(bang_ke_df: pd.DataFrame, dg_case: str) -> pd.DataFrame:
    dg_key = clean_key(dg_case)
    if bang_ke_df.empty or bang_ke_df.shape[1] < 1:
        return bang_ke_df.iloc[0:0].copy()
    mask = bang_ke_df.iloc[:, 0].map(clean_key) == dg_key
    return bang_ke_df.loc[mask].copy()


def resolve_bang_ke_npl_column_indices(
    df: pd.DataFrame,
) -> tuple[int | None, int | None, int | None, int | None, int | None]:
    """Mã NPL, Tên NPL, Mô tả, ĐVT (cột N), và cột P (Số lượng)."""
    n = df.shape[1]
    keys = [clean_key(str(c)) for c in df.columns]

    def pick(pred: object) -> int | None:
        for i, k in enumerate(keys):
            if pred(k):
                return i
        return None

    ma_i = pick(lambda k: "manpl" in k or "macv" in k or "mavtdinh" in k or "mavldm" in k)
    ten_i = pick(lambda k: "tennpl" in k or "tencv" in k or "tenvtdinh" in k)
    mota_i = pick(lambda k: "mota" in k and "sanpham" not in k)
    if mota_i is None:
        mota_i = pick(lambda k: "diendai" in k)
    dvt_i = 13 if n > 13 else None
    p_i = 15 if n > 15 else None
    if ma_i is None and n > 2:
        ma_i = 2
    if ten_i is None and n > 4:
        ten_i = 4
    if mota_i is None and n > 10:
        mota_i = 10
    return ma_i, ten_i, mota_i, dvt_i, p_i


def format_bk_cell(val: object) -> str:
    if pd.isna(val):
        return ""
    return clean_text(val)


def build_npl_rows_from_bk_subset(sub: pd.DataFrame) -> tuple[list[tuple[str, str, str, str, str]], str | None]:
    """Từ các dòng Bảng kê đã lọc theo DG — không đọc file."""
    if sub.empty:
        return [], "Không có dòng Bảng kê cho DG này."
    ma_i, ten_i, mota_i, dvt_i, p_i = resolve_bang_ke_npl_column_indices(sub)
    rows: list[tuple[str, str, str, str, str]] = []
    for _, r in sub.iterrows():

        def cell(idx: int | None) -> str:
            if idx is None or idx >= len(r):
                return ""
            return format_bk_cell(r.iloc[idx])

        rows.append((cell(ma_i), cell(ten_i), cell(mota_i), cell(dvt_i), cell(p_i)))
    return rows, None


def bang_ke_npl_rows_for_display(
    bang_ke_path: str, dg_case: str
) -> tuple[list[tuple[str, str, str, str, str]], str | None]:
    """
    Trả về (danh_sách_dòng, lỗi).
    Mỗi dòng: Mã NPL, Tên NPL, Mô tả, ĐVT (cột N), Số lượng (cột P).
    """
    if not bang_ke_path or not Path(bang_ke_path).exists():
        return [], "Không có file Bảng kê hoặc file không tồn tại."
    try:
        header_row = find_bang_ke_header_row(bang_ke_path)
        bk = pd.read_excel(bang_ke_path, sheet_name=0, header=header_row)
    except Exception as exc:
        return [], f"Không đọc được Bảng kê: {exc}"
    sub = filter_bang_ke_rows_for_dg(bk, dg_case)
    return build_npl_rows_from_bk_subset(sub)


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def format_size_part_rounded(value: object) -> str:
    num = to_number(value)
    if num is not None:
        return str(int(round(num)))
    return clean_text(value)


def format_status_display(status_core: str, reason: str) -> str:
    core = clean_text(status_core)
    r = clean_text(reason)
    if r:
        return f"{core} — Lý do: {r}"
    return core


def parse_status_core_from_display(display: str) -> str:
    s = clean_text(display)
    if not s:
        return ""
    if s.startswith(USER_CONFIRM_OK):
        return USER_CONFIRM_OK
    if s.startswith(USER_CONFIRM_BAD):
        return USER_CONFIRM_BAD
    if s.startswith("Đúng"):
        return "Đúng"
    if s.startswith("Lệch"):
        return "Lệch"
    return s


def ol_production_no_from_groups(order_group: pd.DataFrame, shipped_group: pd.DataFrame) -> str:
    if not order_group.empty and order_group.shape[1] > OL_COL_PRODUCTION_NO:
        v = mode_value(order_group.iloc[:, OL_COL_PRODUCTION_NO])
        if clean_text(v):
            return clean_text(v)
    if not shipped_group.empty and shipped_group.shape[1] > OL_COL_PRODUCTION_NO:
        return clean_text(mode_value(shipped_group.iloc[:, OL_COL_PRODUCTION_NO]))
    return ""


def ol_ship_date_display_and_sort_days(base_group: pd.DataFrame) -> tuple[str, float]:
    if base_group.empty or base_group.shape[1] <= OL_COL_SHIP_DATE:
        return "", 1e15
    ship_ts = pd.to_datetime(base_group.iloc[:, OL_COL_SHIP_DATE], errors="coerce")
    today_start = pd.Timestamp.now().normalize()
    future_ship = ship_ts[ship_ts >= today_start]
    sort_days = (
        float((future_ship.min() - today_start).total_seconds()) if not future_ship.empty else 1e15
    )
    if not future_ship.empty:
        disp = future_ship.min().strftime("%Y-%m-%d")
    elif ship_ts.notna().any():
        disp = pd.Timestamp(ship_ts.max()).strftime("%Y-%m-%d")
    else:
        disp = ""
    return disp, sort_days


def machine_case_conclusion_from_rows(sub: pd.DataFrame) -> str:
    if sub.empty:
        return "Đúng"
    fn = sub["field_name"].astype(str)
    core_mask = ~fn.str.startswith(RULE_CHECK_FIELD_PREFIX, na=False)
    rule_mask = fn.str.startswith(RULE_CHECK_FIELD_PREFIX, na=False)
    if rule_mask.any() and (sub.loc[rule_mask, "status_core"] == "Lệch").any():
        return CASE_CONCLUSION_RULE_BAD
    if core_mask.any() and (sub.loc[core_mask, "status_core"] == "Lệch").any():
        return "Lệch"
    return "Đúng"


def display_case_conclusion(machine: str, user_conclusion: str) -> str:
    u = clean_text(user_conclusion)
    if u == USER_CONFIRM_OK:
        return USER_CONFIRM_OK
    if u == USER_CONFIRM_BAD:
        return USER_CONFIRM_BAD
    return machine if machine in ("Đúng", "Lệch", CASE_CONCLUSION_RULE_BAD) else "Lệch"


def case_row_tag(conclusion: str) -> str:
    if conclusion == USER_CONFIRM_OK:
        return "c_ok"
    if conclusion == USER_CONFIRM_BAD:
        return "c_bad"
    if conclusion == "Đúng":
        return "m_ok"
    if conclusion == "Lệch":
        return "m_bad"
    if conclusion == CASE_CONCLUSION_RULE_BAD:
        return "rule_bad"
    return "m_bad"


def display_case_note(machine: str, case_note: str) -> str:
    if clean_text(case_note):
        return clean_text(case_note)
    return machine


def group_has_future_ship_date(group: pd.DataFrame, ship_col: int = OL_COL_SHIP_DATE) -> bool:
    today_start = pd.Timestamp.now().normalize()
    for v in group.iloc[:, ship_col]:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.notna(ts) and ts >= today_start:
            return True
    return False


def annotate_bang_ke_for_fast_lookup(bang_ke_df: pd.DataFrame) -> pd.DataFrame:
    """Thêm cột phụ (vector) để tránh apply(axis=1) lặp lại theo từng DG — giảm lag Run All."""
    out = bang_ke_df
    if out.shape[1] < 12:
        return out
    col9 = out.iloc[:, 9]
    col10 = out.iloc[:, 10]
    col11 = out.iloc[:, 11]
    k9 = col9.map(lambda v: clean_key(v))
    k10 = col10.map(lambda v: clean_key(v))
    k11 = col11.map(lambda v: clean_key(v))
    out = out.copy()
    out["_dg_key"] = out.iloc[:, 0].map(clean_key)
    out["_pallet948"] = k9.map(lambda s: "948pallet" in s)
    out["_carton_row"] = (
        k10.str.contains("cartonbox", regex=False, na=False)
        | k11.str.contains("cartonbox", regex=False, na=False)
        | k10.str.contains("carton", regex=False, na=False)
        | k11.str.contains("carton", regex=False, na=False)
    )
    out["_fabric"] = k11.map(lambda s: "vai" in s)
    return out


def extract_order_like_metrics(group: pd.DataFrame) -> dict[str, object]:
    if group.empty:
        return {
            "order_no": "",
            "qty_total": 0.0,
            "logo": "",
            "logo_type": "NO LOGO",
            "ma_sp": "",
            "ten_sp": "",
            "color_k": "",
            "color_s": "",
            "size_t": "",
            "size_u": "",
            "carton_qty": 0.0,
        }
    return {
        "order_no": mode_value(group.iloc[:, 1]),
        "qty_total": sum(v for v in group.iloc[:, 6].apply(to_number) if v is not None),
        "logo": normalize_logo(mode_value(group.iloc[:, 9])),
        "logo_type": logo_type_from_order_logo_text(mode_value(group.iloc[:, 17])),
        "ma_sp": mode_value(group.iloc[:, 7]),
        "ten_sp": mode_value(group.iloc[:, 8]),
        "color_k": mode_value(group.iloc[:, 10]),
        "color_s": mode_value(group.iloc[:, 18]),
        "size_t": mode_value(group.iloc[:, 19]),
        "size_u": mode_value(group.iloc[:, 20]),
        "carton_qty": sum(v for v in group.iloc[:, 22].apply(to_number) if v is not None),
    }


class OrderlistCheckerApp:
    def __init__(self, root: tk.Tk, back_to_launcher: callable | None = None):
        init_db()
        self.root = root
        self.back_to_launcher = back_to_launcher
        self.root.title("ORDERLIST EMG Checker")
        self.root.geometry("1300x760")

        self.order_file_var = tk.StringVar()
        self.shipped_file_var = tk.StringVar()
        self.bang_ke_file_var = tk.StringVar()
        self.dg_case_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Sẵn sàng.")
        self.clock_var = tk.StringVar(value="")
        self.date_from_var = tk.StringVar(value="")
        self.date_to_var = tk.StringVar(value="")

        self.last_result_df: pd.DataFrame | None = None
        self.last_run_type: str = "all"
        self.last_target_dg: str | None = None
        self.last_run_customer_filter: str = CUSTOMER_FILTER_EMG
        self.current_run_id: int | None = None
        self.current_view_mode = "summary"
        self.run_case_meta_by_dg: dict[str, dict[str, str]] = {}
        self.history_case_meta_by_dg: dict[str, dict[str, str]] = {}
        self.history_current_run_type: str = ""
        self.history_current_run_id: int | None = None
        self.history_selected_bang_ke_file: str = ""
        self.history_last_df: pd.DataFrame | None = None
        self.detail_win: tk.Toplevel | None = None
        self.detail_tree: ttk.Treeview | None = None
        self.detail_dg_case: str | None = None

        self.check_filter_o_var = tk.StringVar()
        self.check_filter_ma_var = tk.StringVar()
        self.check_filter_status_var = tk.StringVar(value="Tất cả")
        self.customer_filter_var = tk.StringVar(value=CUSTOMER_FILTER_EMG)
        self.customer_filter_values: list[str] = [
            CUSTOMER_FILTER_EMG,
            CUSTOMER_FILTER_NON_EMG,
            CUSTOMER_FILTER_ALL,
        ]
        self.hist_filter_o_var = tk.StringVar()
        self.hist_filter_ma_var = tk.StringVar()
        self.hist_filter_status_var = tk.StringVar(value="Tất cả")
        self.hist_filter_run_label_var = tk.StringVar()
        self.rules_filter_dg_var = tk.StringVar()

        self.config = load_config()
        self.page_size = 50
        self.check_page = 1
        self.check_total_pages = 1
        self.check_view_rows: list[dict] = []
        self._run_progress_total = 0
        self._run_progress_io_active = False
        self._bang_ke_max_cached_files = 3
        self._bang_ke_df_by_path: OrderedDict[str, pd.DataFrame] = OrderedDict()
        self.color_pairs: list[dict[str, str]] = load_color_pairs_from_config(self.config)
        self.color_alias_by_code: dict[str, set[str]] = build_color_alias_lookup(self.color_pairs)
        self.rules_page_size = 50
        self.rules_page = 1
        self.rules_total_pages = 1
        self.rules_all_rows: list[tuple] = []

        self._build_ui()
        self._bind_live_filters()
        self._load_last_paths()
        self._set_default_dates()
        self.refresh_history_runs()
        self.refresh_rules_tab()
        self._start_clock()

    def _build_ui(self) -> None:
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        self.tab_check = ttk.Frame(self.notebook, padding=0)
        self.tab_history = ttk.Frame(self.notebook, padding=0)
        self.tab_rules = ttk.Frame(self.notebook, padding=0)
        self.tab_colors = ttk.Frame(self.notebook, padding=0)
        self.tab_data = ttk.Frame(self.notebook, padding=0)
        self.notebook.add(self.tab_check, text="Check")
        self.notebook.add(self.tab_history, text="Lịch sử")
        self.notebook.add(self.tab_rules, text="Quy tắc")
        self.notebook.add(self.tab_colors, text="Thiết lập màu")
        self.notebook.add(self.tab_data, text="Data")

        top = ttk.Frame(self.tab_check, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="File ORDER LIST (sheet 1):").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(top, textvariable=self.order_file_var, width=100).grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=4)
        ttk.Button(top, text="Chọn file", command=self.choose_order_file).grid(row=0, column=2, pady=4)

        ttk.Label(top, text="File Bảng Kê định mức:").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(top, textvariable=self.bang_ke_file_var, width=100).grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=4)
        ttk.Button(top, text="Chọn file", command=self.choose_bang_ke_file).grid(row=1, column=2, pady=4)

        ttk.Label(top, text="File SHIPPED LIST (sheet 1):").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=4)
        ttk.Entry(top, textvariable=self.shipped_file_var, width=100).grid(row=2, column=1, sticky="ew", padx=(0, 8), pady=4)
        ttk.Button(top, text="Chọn file", command=self.choose_shipped_file).grid(row=2, column=2, pady=4)

        ttk.Label(top, text="Run đơn theo DG Case No:").grid(row=3, column=0, sticky="w", padx=(0, 8), pady=(10, 4))
        ttk.Entry(top, textvariable=self.dg_case_var, width=32).grid(row=3, column=1, sticky="w", pady=(10, 4))
        action = ttk.Frame(top)
        action.grid(row=3, column=2, sticky="e", pady=(10, 4))
        ttk.Button(action, text="Run All", command=self.run_all).pack(side="left")
        ttk.Button(action, text="Run One", command=self.run_one).pack(side="left", padx=(6, 0))
        ttk.Label(top, text="Khách hàng (cột F) cho Run All:").grid(row=4, column=0, sticky="w", padx=(0, 8), pady=(0, 4))
        self.customer_filter_combo = ttk.Combobox(
            top,
            textvariable=self.customer_filter_var,
            values=self.customer_filter_values,
            width=30,
            state="readonly",
        )
        self.customer_filter_combo.grid(row=4, column=1, sticky="w", pady=(0, 4))

        top.columnconfigure(1, weight=1)

        status_wrap = ttk.Frame(self.tab_check, padding=(10, 0, 10, 6))
        status_wrap.pack(fill="x")
        ttk.Label(
            status_wrap,
            textvariable=self.status_var,
            foreground="#1f4e79",
            wraplength=720,
        ).pack(side="left", fill="x", expand=True)
        ttk.Label(status_wrap, textvariable=self.clock_var, foreground="#666666").pack(side="right", padx=(8, 0))

        # Hàng riêng full width: dễ thấy hơn (trước đây nằm cạnh status dễ bị "nuốt").
        self.progress_row = ttk.Frame(self.tab_check, padding=(10, 0, 10, 8))
        self.progress_row.pack(fill="x")
        self.run_progress = ttk.Progressbar(self.progress_row, mode="determinate", length=400, maximum=100)
        self.run_progress.pack(fill="x", expand=True)
        self.run_progress.pack_forget()

        filter_check = ttk.Frame(self.tab_check, padding=(10, 0, 10, 4))
        filter_check.pack(fill="x")
        ttk.Label(filter_check, text="Lọc số O (DG):").pack(side="left")
        ttk.Entry(filter_check, textvariable=self.check_filter_o_var, width=22).pack(side="left", padx=(6, 14))
        ttk.Label(filter_check, text="Lọc mã:").pack(side="left")
        ttk.Entry(filter_check, textvariable=self.check_filter_ma_var, width=18).pack(side="left", padx=(6, 14))
        ttk.Label(filter_check, text="Trạng thái:").pack(side="left")
        ttk.Combobox(
            filter_check,
            textvariable=self.check_filter_status_var,
            values=CHECK_STATUS_FILTER_VALUES,
            width=22,
            state="readonly",
        ).pack(side="left", padx=(6, 10))

        columns = ("dg_case_no", "production_no", "ship_date", "ket_qua", "ket_luan", "ghi_chu")
        headings = {
            "dg_case_no": "DG Case",
            "production_no": "Production No.",
            "ship_date": "Ship date",
            "ket_qua": "Kết quả check",
            "ket_luan": "Kết luận",
            "ghi_chu": "Ghi chú",
        }
        frame = ttk.Frame(self.tab_check, padding=(10, 0, 10, 10))
        frame.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(frame, columns=columns, show="headings")
        for col, width in [
            ("dg_case_no", 140),
            ("production_no", 120),
            ("ship_date", 100),
            ("ket_qua", 120),
            ("ket_luan", 160),
            ("ghi_chu", 280),
        ]:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=width, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.tag_configure("c_ok", background="#a5d6a7")
        self.tree.tag_configure("m_ok", background="#b8d4f0")
        self.tree.tag_configure("m_bad", background="#fff59d")
        self.tree.tag_configure("rule_bad", background="#ffcc80")
        self.tree.tag_configure("c_bad", background="#ef9a9a")
        self.tree.bind("<Double-1>", self.on_tree_double_click)

        action_bottom = ttk.Frame(self.tab_check, padding=(10, 0, 10, 10))
        action_bottom.pack(fill="x")
        ttk.Button(
            action_bottom,
            text="Thiết lập Quy tắc",
            command=self.open_rule_setup_from_check,
        ).pack(side="left")
        if self.back_to_launcher is not None:
            ttk.Button(
                action_bottom,
                text="Back về Launcher",
                command=self.go_back_to_launcher,
            ).pack(side="left", padx=(8, 0))
        pager = ttk.Frame(action_bottom)
        pager.pack(side="right")
        ttk.Button(pager, text="<< Trước", command=self.prev_check_page).pack(side="left")
        self.check_page_label = ttk.Label(pager, text="Trang 1/1", padding=(8, 0))
        self.check_page_label.pack(side="left")
        ttk.Button(pager, text="Sau >>", command=self.next_check_page).pack(side="left")

        self._build_history_tab()
        self._build_rules_tab()
        self._build_color_settings_tab()
        self._build_data_tab()

    def _load_color_pairs_for_ui(self) -> list[dict[str, str]]:
        return load_color_pairs_from_config(self.config)

    def _build_color_settings_tab(self) -> None:
        outer = ttk.Frame(self.tab_colors, padding=(10, 10, 10, 10))
        outer.pack(fill="both", expand=True)
        ttk.Label(
            outer,
            text="Bảng thiết lập mã màu dùng để đối chiếu màu trên OL/SHIPPED.",
            wraplength=960,
        ).pack(anchor="w", pady=(0, 6))
        ttk.Label(
            outer,
            text="Có thể thêm/sửa/xóa tự do. Một mã màu có thể khai báo nhiều tên (ví dụ: Gray/Grey, Blue/Marine Blue).",
            foreground="#555555",
            wraplength=960,
        ).pack(anchor="w", pady=(0, 10))

        table_wrap = ttk.Frame(outer)
        table_wrap.pack(fill="both", expand=True)
        cols = ("code", "name")
        tree = ttk.Treeview(table_wrap, columns=cols, show="headings", height=14)
        tree.heading("code", text="Mã màu")
        tree.heading("name", text="Tên màu")
        tree.column("code", width=140, anchor="center")
        tree.column("name", width=320, anchor="w")
        tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(table_wrap, orient="vertical", command=tree.yview)
        sb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=sb.set)

        input_row = ttk.Frame(outer)
        input_row.pack(fill="x", pady=(10, 0))
        ttk.Label(input_row, text="Mã màu:").pack(side="left")
        code_var = tk.StringVar()
        name_var = tk.StringVar()
        ttk.Entry(input_row, textvariable=code_var, width=16).pack(side="left", padx=(6, 12))
        ttk.Label(input_row, text="Tên màu:").pack(side="left")
        ttk.Entry(input_row, textvariable=name_var, width=38).pack(side="left", padx=(6, 12))

        def _fill_tree(rows: list[dict[str, str]]) -> None:
            for iid in tree.get_children():
                tree.delete(iid)
            for item in rows:
                tree.insert("", "end", values=(item["code"], item["name"]))

        def _read_rows_from_tree() -> list[dict[str, str]]:
            rows: list[dict[str, str]] = []
            for iid in tree.get_children():
                vals = tree.item(iid, "values")
                if not vals:
                    continue
                code = clean_text(vals[0] if len(vals) > 0 else "")
                name = clean_text(vals[1] if len(vals) > 1 else "")
                if code and name:
                    rows.append({"code": code, "name": name})
            return dedupe_color_pairs(rows)

        def _on_tree_select(_: object | None = None) -> None:
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], "values")
            code_var.set(clean_text(vals[0] if len(vals) > 0 else ""))
            name_var.set(clean_text(vals[1] if len(vals) > 1 else ""))

        tree.bind("<<TreeviewSelect>>", _on_tree_select)

        def _add_or_update_selected() -> None:
            code = clean_text(code_var.get())
            name = clean_text(name_var.get())
            if not code or not name:
                messagebox.showwarning("Thiếu dữ liệu", "Nhập đủ Mã màu và Tên màu.")
                return
            sel = tree.selection()
            if sel:
                tree.item(sel[0], values=(code, name))
            else:
                tree.insert("", "end", values=(code, name))
            rows = _read_rows_from_tree()
            _fill_tree(rows)
            code_var.set("")
            name_var.set("")

        def _delete_selected() -> None:
            sel = tree.selection()
            if not sel:
                messagebox.showinfo("Thiếu dòng chọn", "Chọn một dòng để xóa.")
                return
            tree.delete(sel[0])

        def _reset_defaults() -> None:
            _fill_tree([{"code": c, "name": n} for c, n in DEFAULT_COLOR_CODE_PAIRS])

        def _save_color_settings() -> None:
            rows = _read_rows_from_tree()
            if not rows:
                messagebox.showwarning("Thiếu dữ liệu", "Bảng màu đang trống.")
                return
            self.config[COLOR_MAP_CONFIG_KEY] = rows
            save_config(self.config)
            self.color_pairs = load_color_pairs_from_config(self.config)
            self.color_alias_by_code = build_color_alias_lookup(self.color_pairs)
            self._refresh_check_tab_current_data("mau")
            self.status_var.set("Đã lưu thiết lập màu sắc.")
            messagebox.showinfo("Thiết lập màu", f"Đã lưu {len(rows)} dòng mapping màu.")

        btn_row = ttk.Frame(outer)
        btn_row.pack(fill="x", pady=(10, 0))
        ttk.Button(btn_row, text="Thêm / Cập nhật dòng chọn", command=_add_or_update_selected).pack(side="left")
        ttk.Button(btn_row, text="Xóa dòng chọn", command=_delete_selected).pack(side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Khôi phục mặc định", command=_reset_defaults).pack(side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Lưu thiết lập màu", command=_save_color_settings).pack(side="left", padx=(18, 0))

        _fill_tree(self._load_color_pairs_for_ui())

    def _build_rules_tab(self) -> None:
        top = ttk.Frame(self.tab_rules, padding=(10, 10, 10, 6))
        top.pack(fill="x")
        ttk.Label(
            top,
            text="Quy tắc theo DG Case — khi kích hoạt, mỗi Run sẽ thêm bước so OL với giá trị kỳ vọng. Ô để trống = bỏ qua tiêu chí đó.",
            wraplength=960,
        ).pack(anchor="w")
        filter_row = ttk.Frame(self.tab_rules, padding=(10, 0, 10, 6))
        filter_row.pack(fill="x")
        ttk.Label(filter_row, text="Lọc DG Case:").pack(side="left")
        ttk.Entry(filter_row, textvariable=self.rules_filter_dg_var, width=26).pack(side="left", padx=(6, 10))
        pager = ttk.Frame(filter_row)
        pager.pack(side="right")
        ttk.Button(pager, text="<< Trước", command=self.prev_rules_page).pack(side="left")
        self.rules_page_label = ttk.Label(pager, text="Trang 1/1", padding=(8, 0))
        self.rules_page_label.pack(side="left")
        ttk.Button(pager, text="Sau >>", command=self.next_rules_page).pack(side="left")
        wrap = ttk.Frame(self.tab_rules, padding=(10, 0, 10, 6))
        wrap.pack(fill="both", expand=True)
        rcols = ("dg_case", "active", "n_crit", "created_at", "updated_at")
        self.rules_tree = ttk.Treeview(wrap, columns=rcols, show="headings", height=18)
        rh = {
            "dg_case": "DG Case",
            "active": "Kích hoạt",
            "n_crit": "Số tiêu chí đặt",
            "created_at": "Ngày lập",
            "updated_at": "Cập nhật",
        }
        for c, w, st in [
            ("dg_case", 160, False),
            ("active", 72, False),
            ("n_crit", 100, False),
            ("created_at", 150, False),
            ("updated_at", 150, True),
        ]:
            self.rules_tree.heading(c, text=rh[c])
            self.rules_tree.column(c, width=w, anchor="center", stretch=st)
        self.rules_tree.pack(side="left", fill="both", expand=True)
        rsb = ttk.Scrollbar(wrap, orient="vertical", command=self.rules_tree.yview)
        rsb.pack(side="right", fill="y")
        self.rules_tree.configure(yscrollcommand=rsb.set)
        btnf = ttk.Frame(self.tab_rules, padding=(10, 0, 10, 10))
        btnf.pack(fill="x")
        ttk.Button(btnf, text="Thiết lập quy tắc (nhập DG)", command=self.open_rule_setup_new_from_rules_tab).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(btnf, text="Sửa quy tắc", command=self.edit_rule_from_rules_tab).pack(side="left", padx=(0, 8))
        ttk.Button(btnf, text="UnActive", command=self.deactivate_rule_from_rules_tab).pack(side="left", padx=(0, 8))
        ttk.Button(btnf, text="Kích hoạt lại", command=self.activate_rule_from_rules_tab).pack(side="left", padx=(0, 8))
        ttk.Button(btnf, text="Làm mới danh sách", command=self.refresh_rules_tab).pack(side="left", padx=(16, 0))

    def refresh_rules_tab(self) -> None:
        if not hasattr(self, "rules_tree"):
            return
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        rows = cur.execute(
            """
            SELECT
                s.dg_case_no,
                s.is_active,
                (SELECT COUNT(*) FROM dg_rule_criteria c
                 WHERE c.rule_set_id = s.id AND trim(COALESCE(c.expected_value, '')) != ''),
                COALESCE(s.created_at, ''),
                COALESCE(s.updated_at, '')
            FROM dg_rule_sets s
            ORDER BY s.updated_at DESC, s.dg_case_no
            """
        ).fetchall()
        conn.close()
        built: list[tuple] = []
        for dg, active, n, cr, up in rows:
            act_txt = "Có" if int(active) == 1 else "Không"
            built.append((dg, act_txt, int(n or 0), clean_text(cr), clean_text(up)))
        self.rules_all_rows = built
        self.rules_page = 1
        self._apply_rules_filter_and_render(reset_page=False)

    def _apply_rules_filter_and_render(self, reset_page: bool = True) -> None:
        if not hasattr(self, "rules_tree"):
            return
        key = clean_key(self.rules_filter_dg_var.get())
        rows = self.rules_all_rows
        if key:
            rows = [r for r in rows if key in clean_key(r[0])]
        if reset_page:
            self.rules_page = 1
        total = len(rows)
        self.rules_total_pages = max(1, (total + self.rules_page_size - 1) // self.rules_page_size)
        if self.rules_page > self.rules_total_pages:
            self.rules_page = self.rules_total_pages
        if self.rules_page < 1:
            self.rules_page = 1
        for iid in self.rules_tree.get_children():
            self.rules_tree.delete(iid)
        start = (self.rules_page - 1) * self.rules_page_size
        end = start + self.rules_page_size
        for row in rows[start:end]:
            self.rules_tree.insert("", "end", values=row)
        if hasattr(self, "rules_page_label"):
            self.rules_page_label.config(text=f"Trang {self.rules_page}/{self.rules_total_pages}")

    def prev_rules_page(self) -> None:
        if self.rules_page > 1:
            self.rules_page -= 1
            self._apply_rules_filter_and_render(reset_page=False)

    def next_rules_page(self) -> None:
        if self.rules_page < self.rules_total_pages:
            self.rules_page += 1
            self._apply_rules_filter_and_render(reset_page=False)

    def _rules_tab_selected_dg(self) -> str:
        sel = self.rules_tree.selection()
        if not sel:
            return ""
        vals = self.rules_tree.item(sel[0], "values")
        return clean_text(str(vals[0])) if vals else ""

    def edit_rule_from_rules_tab(self) -> None:
        dg = self._rules_tab_selected_dg()
        if not dg:
            messagebox.showinfo("Thông báo", "Chọn một dòng quy tắc (DG Case).")
            return
        self._open_rule_editor_dialog(initial_dg=dg, go_to_tab_on_save=False)

    def open_rule_setup_new_from_rules_tab(self) -> None:
        self._open_rule_editor_dialog(initial_dg="", go_to_tab_on_save=False)

    def deactivate_rule_from_rules_tab(self) -> None:
        dg = self._rules_tab_selected_dg()
        if not dg:
            messagebox.showinfo("Thông báo", "Chọn một dòng quy tắc (DG Case).")
            return
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "UPDATE dg_rule_sets SET is_active = 0, updated_at = CURRENT_TIMESTAMP WHERE dg_case_no = ?",
            (dg,),
        )
        conn.commit()
        conn.close()
        self.refresh_rules_tab()

    def _build_data_tab(self) -> None:
        top = ttk.Frame(self.tab_data, padding=(10, 10, 10, 6))
        top.pack(fill="x")
        ttk.Label(
            top,
            text=(
                "Export: chọn 1 run rồi xuất ra JSON để lưu/chia sẻ. "
                "Import: nạp JSON run đã xuất để thêm vào Lịch sử."
            ),
            wraplength=980,
        ).pack(anchor="w")

        act = ttk.Frame(self.tab_data, padding=(10, 0, 10, 6))
        act.pack(fill="x")
        ttk.Button(act, text="Nạp danh sách run", command=self.refresh_data_runs).pack(side="left")
        ttk.Button(act, text="Xuất run đã chọn", command=self.export_selected_run_data).pack(side="left", padx=(8, 0))
        ttk.Button(act, text="Import run từ file", command=self.import_run_data_file).pack(side="left", padx=(8, 0))

        wrap = ttk.Frame(self.tab_data, padding=(10, 0, 10, 10))
        wrap.pack(fill="both", expand=True)
        cols = ("run_id", "created_at", "run_label", "run_type", "total", "lech", "manual_check")
        self.data_runs_tree = ttk.Treeview(wrap, columns=cols, show="headings", height=16)
        hd = {
            "run_id": "run_id",
            "created_at": "created_at",
            "run_label": "Tên run / ghi chú",
            "run_type": "run_type",
            "total": "total",
            "lech": "lech",
            "manual_check": "Check",
        }
        for c, w, st in [
            ("run_id", 64, False),
            ("created_at", 140, False),
            ("run_label", 220, True),
            ("run_type", 84, False),
            ("total", 80, False),
            ("lech", 80, False),
            ("manual_check", 80, False),
        ]:
            self.data_runs_tree.heading(c, text=hd[c])
            self.data_runs_tree.column(c, width=w, anchor="center", stretch=st)
        self.data_runs_tree.pack(side="left", fill="both", expand=True)
        scr = ttk.Scrollbar(wrap, orient="vertical", command=self.data_runs_tree.yview)
        scr.pack(side="right", fill="y")
        self.data_runs_tree.configure(yscrollcommand=scr.set)
        self.refresh_data_runs()

    def refresh_data_runs(self) -> None:
        if not hasattr(self, "data_runs_tree"):
            return
        for item in self.data_runs_tree.get_children():
            self.data_runs_tree.delete(item)
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        rows = cur.execute(
            """
            SELECT
                r.id,
                r.created_at,
                COALESCE(r.run_label, ''),
                r.run_type,
                COUNT(i.id) AS total_checks,
                SUM(CASE WHEN i.status = 'Lệch' THEN 1 ELSE 0 END) AS bad_checks,
                (SELECT COUNT(*) FROM run_case_meta m WHERE m.run_id = r.id) AS manual_case_checks
            FROM runs r
            LEFT JOIN run_items i ON i.run_id = r.id
            GROUP BY r.id
            ORDER BY r.id DESC
            """
        ).fetchall()
        conn.close()
        for row in rows:
            self.data_runs_tree.insert("", "end", values=row)

    def _selected_data_run_id(self) -> int | None:
        if not hasattr(self, "data_runs_tree"):
            return None
        sel = self.data_runs_tree.selection()
        if not sel:
            return None
        vals = self.data_runs_tree.item(sel[0], "values")
        if not vals:
            return None
        try:
            return int(vals[0])
        except (TypeError, ValueError):
            return None

    def export_selected_run_data(self) -> None:
        run_id = self._selected_data_run_id()
        if run_id is None:
            messagebox.showinfo("Export run", "Chọn 1 run ở bảng Data trước.")
            return
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        run_row = cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        if run_row is None:
            conn.close()
            messagebox.showerror("Export run", "Không tìm thấy run trong DB.")
            return
        items = [dict(r) for r in cur.execute("SELECT * FROM run_items WHERE run_id = ? ORDER BY id", (run_id,)).fetchall()]
        meta = [dict(r) for r in cur.execute("SELECT * FROM run_case_meta WHERE run_id = ? ORDER BY dg_case_no", (run_id,)).fetchall()]
        conn.close()
        payload = {
            "format": "orderlist_emg_checker_run_export_v1",
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "run": dict(run_row),
            "run_items": items,
            "run_case_meta": meta,
        }
        suggested = f"run_{run_id}.json"
        out = filedialog.asksaveasfilename(
            title="Xuất dữ liệu run",
            defaultextension=".json",
            initialfile=suggested,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not out:
            return
        Path(out).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        messagebox.showinfo("Export run", f"Đã xuất run {run_id} ra file:\n{out}")

    def import_run_data_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Import dữ liệu run (JSON)",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception as exc:
            messagebox.showerror("Import run", f"Không đọc được JSON:\n{exc}")
            return
        if clean_text(payload.get("format")) != "orderlist_emg_checker_run_export_v1":
            messagebox.showerror("Import run", "Sai định dạng file export run.")
            return
        run = payload.get("run") or {}
        items = payload.get("run_items") or []
        meta = payload.get("run_case_meta") or []
        if not isinstance(run, dict) or not isinstance(items, list) or not isinstance(meta, list):
            messagebox.showerror("Import run", "Nội dung JSON không hợp lệ.")
            return
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("PRAGMA foreign_keys = ON")
        cur.execute(
            """
            INSERT INTO runs (
                run_type, target_dg, order_file, bang_ke_file, run_label, created_at,
                trace_orderlist_filename, trace_shipped_file_mtime, trace_bang_ke_a6
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                clean_text(run.get("run_type")),
                clean_text(run.get("target_dg")),
                clean_text(run.get("order_file")),
                clean_text(run.get("bang_ke_file")),
                clean_text(run.get("run_label")),
                clean_text(run.get("created_at")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                clean_text(run.get("trace_orderlist_filename")),
                clean_text(run.get("trace_shipped_file_mtime")),
                clean_text(run.get("trace_bang_ke_a6")),
            ),
        )
        new_run_id = int(cur.lastrowid)
        item_rows: list[tuple] = []
        for r in items:
            if not isinstance(r, dict):
                continue
            item_rows.append(
                (
                    new_run_id,
                    clean_text(r.get("dg_case_no")),
                    clean_text(r.get("field_name")),
                    clean_text(r.get("order_value")),
                    clean_text(r.get("shipped_value")),
                    clean_text(r.get("bang_ke_value")),
                    clean_text(r.get("auto_status")) or clean_text(r.get("status")),
                    clean_text(r.get("status")),
                    int(to_number(r.get("is_adjusted")) or 0),
                    clean_text(r.get("adjust_reason")),
                    clean_text(r.get("production_no")),
                    clean_text(r.get("ship_date_display")),
                )
            )
        if item_rows:
            cur.executemany(
                """
                INSERT INTO run_items (
                    run_id, dg_case_no, field_name, order_value, shipped_value, bang_ke_value,
                    auto_status, status, is_adjusted, adjust_reason, production_no, ship_date_display
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                item_rows,
            )
        meta_rows: list[tuple] = []
        for r in meta:
            if not isinstance(r, dict):
                continue
            meta_rows.append(
                (
                    new_run_id,
                    clean_text(r.get("dg_case_no")),
                    clean_text(r.get("user_conclusion")),
                    clean_text(r.get("case_note")),
                    clean_text(r.get("updated_at")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                )
            )
        if meta_rows:
            cur.executemany(
                """
                INSERT INTO run_case_meta (run_id, dg_case_no, user_conclusion, case_note, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                meta_rows,
            )
        conn.commit()
        conn.close()
        self.refresh_history_runs()
        self.refresh_data_runs()
        messagebox.showinfo("Import run", f"Đã import thành công run mới (id={new_run_id}).")

    def activate_rule_from_rules_tab(self) -> None:
        dg = self._rules_tab_selected_dg()
        if not dg:
            messagebox.showinfo("Thông báo", "Chọn một dòng quy tắc (DG Case).")
            return
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "UPDATE dg_rule_sets SET is_active = 1, updated_at = CURRENT_TIMESTAMP WHERE dg_case_no = ?",
            (dg,),
        )
        conn.commit()
        conn.close()
        self.refresh_rules_tab()

    def open_rule_setup_from_check(self) -> None:
        dg = ""
        sel = self.tree.selection()
        if sel:
            dg = clean_text(str(self.tree.item(sel[0], "values")[0]))
        self._open_rule_editor_dialog(initial_dg=dg, go_to_tab_on_save=True)

    def open_rule_setup_from_history_tab(self) -> None:
        dg = ""
        if hasattr(self, "history_items_tree"):
            sel = self.history_items_tree.selection()
            if sel:
                dg = clean_text(str(self.history_items_tree.item(sel[0], "values")[0]))
        self._open_rule_editor_dialog(initial_dg=dg, go_to_tab_on_save=True)

    def _prefill_rule_entries_from_history_df(self, dg: str) -> dict[str, str]:
        out = {k: "" for k in RULE_CRITERIA_FIELDS}
        if not dg or self.history_last_df is None or self.history_last_df.empty:
            return out
        sub = self.history_last_df[
            self.history_last_df["dg_case_no"].astype(str).map(lambda x: clean_key(x) == clean_key(dg))
        ]
        if sub.empty:
            return out
        for fn in RULE_CRITERIA_FIELDS:
            rows = sub[sub["field_name"].astype(str) == fn]
            if rows.empty:
                continue
            v = clean_text(str(rows.iloc[0]["order_value"]))
            if "| Tổng:" in v:
                v = v.split("| Tổng:")[0].strip()
            elif "|" in v and fn.startswith("Số"):
                v = v.split("|")[0].strip()
            out[fn] = v
        return out

    def _prefill_rule_entries_merged(self, dg: str) -> dict[str, str]:
        merged = {k: "" for k in RULE_CRITERIA_FIELDS}
        if not dg:
            return merged
        a = self._prefill_rule_entries_from_last_df(dg)
        b = self._prefill_rule_entries_from_history_df(dg)
        for k in RULE_CRITERIA_FIELDS:
            merged[k] = clean_text(a.get(k, "")) or clean_text(b.get(k, ""))
        return merged

    def _prefill_rule_entries_from_last_df(self, dg: str) -> dict[str, str]:
        out = {k: "" for k in RULE_CRITERIA_FIELDS}
        if self.last_result_df is None or self.last_result_df.empty:
            return out
        sub = self.last_result_df[
            self.last_result_df["dg_case_no"].astype(str).map(lambda x: clean_key(x) == clean_key(dg))
        ]
        if sub.empty:
            return out
        for fn in RULE_CRITERIA_FIELDS:
            rows = sub[sub["field_name"].astype(str) == fn]
            if rows.empty:
                continue
            v = clean_text(str(rows.iloc[0]["order_value"]))
            if "| Tổng:" in v:
                v = v.split("| Tổng:")[0].strip()
            elif "|" in v and fn.startswith("Số"):
                v = v.split("|")[0].strip()
            out[fn] = v
        return out

    def _load_rule_criteria_dict(self, dg: str) -> dict[str, str]:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT id FROM dg_rule_sets WHERE dg_case_no = ?", (dg,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return {}
        rid = row[0]
        cur.execute(
            "SELECT field_name, COALESCE(expected_value, '') FROM dg_rule_criteria WHERE rule_set_id = ?",
            (rid,),
        )
        m = {fn: clean_text(ev) for fn, ev in cur.fetchall()}
        conn.close()
        return m

    def _save_rule_criteria_db(self, dg: str, values_by_field: dict[str, str]) -> None:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO dg_rule_sets (dg_case_no, is_active) VALUES (?, 1)
            ON CONFLICT(dg_case_no) DO UPDATE SET
                is_active = 1,
                updated_at = CURRENT_TIMESTAMP
            """,
            (dg,),
        )
        cur.execute("SELECT id FROM dg_rule_sets WHERE dg_case_no = ?", (dg,))
        rid = int(cur.fetchone()[0])
        cur.execute("DELETE FROM dg_rule_criteria WHERE rule_set_id = ?", (rid,))
        for fn in RULE_CRITERIA_FIELDS:
            ev = clean_text(values_by_field.get(fn, ""))
            if not ev:
                continue
            cur.execute(
                "INSERT INTO dg_rule_criteria (rule_set_id, field_name, expected_value) VALUES (?, ?, ?)",
                (rid, fn, ev),
            )
        conn.commit()
        conn.close()

    def _open_rule_editor_dialog(self, initial_dg: str = "", go_to_tab_on_save: bool = False) -> None:
        dg_seed = clean_text(initial_dg)
        saved = self._load_rule_criteria_dict(dg_seed) if dg_seed else {}
        pref = self._prefill_rule_entries_merged(dg_seed) if dg_seed else {k: "" for k in RULE_CRITERIA_FIELDS}
        win = tk.Toplevel(self.root)
        win.title("Thiết lập quy tắc")
        win.geometry("720x660")
        outer = ttk.Frame(win, padding=12)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text="DG Case", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 4))
        dg_var = tk.StringVar(value=dg_seed)
        ttk.Entry(outer, textvariable=dg_var, width=44).pack(anchor="w", pady=(0, 6))
        ttk.Label(
            outer,
            text="Có thể nhập tay mã DG; nếu đã chọn dòng trên bảng Check/Lịch sử, mã sẽ được điền sẵn. "
            "Gợi ý tiêu chí lấy từ Run Check gần nhất hoặc chi tiết run đang mở ở Lịch sử (cùng DG).",
            wraplength=680,
            foreground="#444444",
        ).pack(anchor="w", pady=(0, 8))
        ttk.Label(
            outer,
            text="Điền giá trị kỳ vọng so với OL khi Run. Để trống = bỏ qua tiêu chí. "
            "Size thùng / pallet: định dạng chấm như OL (VD: 209.48.52, 120.100.h). Lưu = cập nhật thời điểm «Cập nhật».",
            wraplength=680,
            foreground="#444444",
        ).pack(anchor="w", pady=(0, 10))
        canvas = tk.Canvas(outer, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        frm = ttk.Frame(canvas)
        frm.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=frm, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        entries: dict[str, tk.Entry] = {}
        for i, fn in enumerate(RULE_CRITERIA_FIELDS):
            ttk.Label(frm, text=fn + ":").grid(row=i, column=0, sticky="nw", pady=3, padx=(0, 8))
            e = ttk.Entry(frm, width=72)
            val = saved.get(fn) or pref.get(fn, "")
            e.insert(0, val)
            e.grid(row=i, column=1, sticky="ew", pady=3)
            entries[fn] = e
        frm.columnconfigure(1, weight=1)

        def _reload_fields_from_dg() -> None:
            dg_cur = clean_text(dg_var.get())
            if not dg_cur:
                return
            sv = self._load_rule_criteria_dict(dg_cur)
            pr = self._prefill_rule_entries_merged(dg_cur)
            for fn in RULE_CRITERIA_FIELDS:
                ent = entries[fn]
                ent.delete(0, "end")
                v = sv.get(fn) or pr.get(fn, "")
                ent.insert(0, v)

        def _save() -> None:
            dg_final = clean_text(dg_var.get())
            if not dg_final:
                messagebox.showwarning("Thiếu DG Case", "Nhập mã DG Case (ví dụ O-05193-01).")
                return
            vals = {fn: entries[fn].get() for fn in RULE_CRITERIA_FIELDS}
            self._save_rule_criteria_db(dg_final, vals)
            self.refresh_rules_tab()
            self._refresh_check_tab_current_data("quy_tac")
            win.destroy()
            if go_to_tab_on_save:
                try:
                    self.notebook.select(self.notebook.index(self.tab_rules))
                except tk.TclError:
                    pass

        bf = ttk.Frame(outer)
        bf.pack(fill="x", pady=(12, 0))
        ttk.Button(bf, text="Lưu quy tắc", command=_save).pack(side="left", padx=(0, 8))
        ttk.Button(bf, text="Nạp lại theo DG", command=_reload_fields_from_dg).pack(side="left", padx=(0, 8))
        ttk.Button(bf, text="Đóng", command=win.destroy).pack(side="left")

    def _build_quy_tac_records(
        self,
        dg_case_disp: str,
        checks: list[tuple[str, str, str, str]],
        prod_no: str,
        ship_date_display: str,
        sort_days: float,
        order_metrics: dict[str, object],
        ol_missing: bool,
        order_size_xyz: str,
    ) -> list[dict]:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            "SELECT s.id FROM dg_rule_sets s WHERE s.dg_case_no = ? AND s.is_active = 1",
            (clean_text(dg_case_disp),),
        )
        row = cur.fetchone()
        if not row:
            conn.close()
            return []
        rid = row[0]
        cur.execute(
            "SELECT field_name, COALESCE(expected_value, '') FROM dg_rule_criteria WHERE rule_set_id = ?",
            (rid,),
        )
        crit = cur.fetchall()
        conn.close()
        ol_map: dict[str, str] = {t[0]: t[1] for t in checks}
        if "Số thùng (pallet)" not in ol_map:
            ol_map["Số thùng (pallet)"] = format_number(float(order_metrics["carton_qty"]))
        if "Size pallet" not in ol_map:
            ol_map["Size pallet"] = order_size_xyz
        out: list[dict] = []
        for fn, ev in crit:
            exp = clean_text(ev)
            if not exp:
                continue
            if fn not in ol_map:
                continue
            ol_raw = ol_map[fn]
            st, adj = rule_compare_ol_expected(fn, ol_raw, exp, ol_missing)
            ol_show = "Hàng đã đi, OL không có dữ liệu" if ol_missing else clean_text(ol_raw)
            out.append(
                {
                    "dg_case_no": dg_case_disp,
                    "field_name": f"{RULE_CHECK_FIELD_PREFIX}{fn}",
                    "order_value": ol_show,
                    "shipped_value": "",
                    "bang_ke_value": f"Kỳ vọng (quy tắc): {exp}",
                    "auto_status": st,
                    "status_core": st,
                    "adjust_reason": adj,
                    "production_no": prod_no,
                    "ship_date_display": ship_date_display,
                    "_sort_days": sort_days,
                }
            )
        return out

    def apply_check_filters(self) -> None:
        if self.last_result_df is None or self.last_result_df.empty:
            messagebox.showinfo("Thông báo", "Chưa có dữ liệu. Hãy Run trước.")
            return
        self.render_check_main_table()

    def _bind_live_filters(self) -> None:
        def _on_check_filter_change(*_args: object) -> None:
            if self.last_result_df is None or self.last_result_df.empty:
                return
            self.render_check_main_table()

        def _on_history_item_filter_change(*_args: object) -> None:
            if self.history_last_df is None or self.history_last_df.empty:
                return
            self._render_history_items_from_df(self.history_last_df, self.history_current_run_type or "one")

        def _on_history_run_filter_change(*_args: object) -> None:
            if not hasattr(self, "history_runs_tree"):
                return
            self.refresh_history_runs()

        def _on_rules_filter_change(*_args: object) -> None:
            self._apply_rules_filter_and_render(reset_page=True)

        self.check_filter_o_var.trace_add("write", _on_check_filter_change)
        self.check_filter_ma_var.trace_add("write", _on_check_filter_change)
        self.check_filter_status_var.trace_add("write", _on_check_filter_change)

        self.hist_filter_o_var.trace_add("write", _on_history_item_filter_change)
        self.hist_filter_ma_var.trace_add("write", _on_history_item_filter_change)
        self.hist_filter_status_var.trace_add("write", _on_history_item_filter_change)

        self.hist_filter_run_label_var.trace_add("write", _on_history_run_filter_change)
        self.date_from_var.trace_add("write", _on_history_run_filter_change)
        self.date_to_var.trace_add("write", _on_history_run_filter_change)
        self.rules_filter_dg_var.trace_add("write", _on_rules_filter_change)

    def prev_check_page(self) -> None:
        if self.check_page > 1:
            self.check_page -= 1
            self._render_check_current_page()

    def next_check_page(self) -> None:
        if self.check_page < self.check_total_pages:
            self.check_page += 1
            self._render_check_current_page()

    def go_back_to_launcher(self) -> None:
        if self.back_to_launcher is None:
            return
        self.root.destroy()
        self.back_to_launcher()

    def _run_progress_io_begin(self) -> None:
        """Giai đoạn đọc Excel nặng — indeterminate để user thấy app không treo."""
        self._run_progress_io_active = True
        self.run_progress.stop()
        self.run_progress.pack_forget()
        self.run_progress.configure(mode="indeterminate", maximum=100)
        self.run_progress.pack(fill="x", expand=True)
        self.run_progress.start(14)
        self.root.update_idletasks()
        self.root.update()

    def _run_progress_io_end(self) -> None:
        if not self._run_progress_io_active:
            return
        self._run_progress_io_active = False
        self.run_progress.stop()
        self.run_progress.pack_forget()
        self.run_progress.configure(mode="determinate", maximum=100, value=0)
        self.root.update_idletasks()

    def _run_progress_start(self, total: int) -> None:
        self._run_progress_total = max(0, int(total))
        if self._run_progress_total <= 0:
            return
        self.run_progress.stop()
        self.run_progress.pack_forget()
        self.run_progress.configure(mode="determinate", maximum=self._run_progress_total, value=0)
        self.run_progress.pack(fill="x", expand=True)
        self.root.update_idletasks()
        self.root.update()

    def _run_progress_update(self, done: int) -> None:
        if self._run_progress_total <= 0:
            return
        self.run_progress.configure(value=min(int(done), self._run_progress_total))
        self.root.update_idletasks()
        self.root.update()

    def _run_progress_done(self) -> None:
        try:
            self.run_progress.stop()
        except tk.TclError:
            pass
        self._run_progress_io_active = False
        self._run_progress_total = 0
        self.run_progress.pack_forget()
        self.run_progress.configure(mode="determinate", maximum=100, value=0)
        self.root.update_idletasks()

    def _set_check_view_rows(self, rows: list[dict]) -> None:
        self.check_view_rows = rows
        total = len(rows)
        self.check_total_pages = max(1, (total + self.page_size - 1) // self.page_size)
        self.check_page = min(self.check_page, self.check_total_pages)
        if self.check_page < 1:
            self.check_page = 1
        self._render_check_current_page()

    def _render_check_current_page(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)
        if not self.check_view_rows:
            self.check_page_label.config(text="Trang 1/1")
            return
        start = (self.check_page - 1) * self.page_size
        end = start + self.page_size
        for row in self.check_view_rows[start:end]:
            self.tree.insert("", "end", values=row["values"], tags=(row["tag"],))
        self.check_page_label.config(text=f"Trang {self.check_page}/{self.check_total_pages}")

    def _filter_result_df_by_o(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        o = self.check_filter_o_var.get().strip()
        if not o:
            return df
        key = clean_key(o)
        return df[
            df["dg_case_no"].astype(str).apply(lambda dg: key == clean_key(dg) or key in clean_key(dg))
        ]

    def _filter_hist_df_by_o(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        o = self.hist_filter_o_var.get().strip()
        if not o:
            return df
        key = clean_key(o)
        return df[df["dg_case_no"].astype(str).apply(lambda dg: key == clean_key(dg) or key in clean_key(dg))]

    def _summary_rows_from_df(self, view_df: pd.DataFrame, meta_by_dg: dict[str, dict[str, str]]) -> list[dict]:
        if view_df.empty:
            return []
        rows: list[dict] = []
        for dg, sub in view_df.groupby("dg_case_no", sort=False):
            dg_s = clean_text(str(dg))
            if not dg_s:
                continue
            total = int(len(sub))
            ok = int((sub["status_core"] == "Đúng").sum())
            machine = machine_case_conclusion_from_rows(sub)
            meta = meta_by_dg.get(dg_s, {})
            user_c = clean_text(meta.get("user_conclusion", ""))
            case_note = clean_text(meta.get("case_note", ""))
            ket_luan = display_case_conclusion(machine, user_c)
            ghi_chu = display_case_note(machine, case_note)
            prod = ""
            ship = ""
            if "production_no" in sub.columns:
                for x in sub["production_no"].tolist():
                    if clean_text(x):
                        prod = clean_text(x)
                        break
            if "ship_date_display" in sub.columns:
                for x in sub["ship_date_display"].tolist():
                    if clean_text(x):
                        ship = clean_text(x)
                        break
            sort_days = 1e15
            if "_sort_days" in sub.columns and not sub.empty:
                mn = pd.to_numeric(sub["_sort_days"], errors="coerce").min()
                sort_days = float(mn) if pd.notna(mn) else 1e15
            elif ship:
                ship_dt = pd.to_datetime(ship, errors="coerce")
                if pd.notna(ship_dt):
                    sort_days = float((ship_dt.normalize() - pd.Timestamp.now().normalize()).total_seconds())
            rows.append(
                {
                    "dg": dg_s,
                    "production_no": prod,
                    "ship_date": ship,
                    "ket_qua": f"{ok}/{total}",
                    "ket_luan": ket_luan,
                    "ghi_chu": ghi_chu,
                    "tag": case_row_tag(ket_luan),
                    "sort_days": sort_days,
                }
            )
        rows.sort(key=lambda r: (r["sort_days"], r["dg"]))
        return rows

    def _filter_summary_rows_by_status(self, rows: list[dict], status_var: tk.StringVar) -> list[dict]:
        st = status_var.get().strip()
        if not st or st == "Tất cả":
            return rows
        return [r for r in rows if r["ket_luan"] == st]

    def _filter_summary_rows_by_production(self, rows: list[dict], ma_var: tk.StringVar) -> list[dict]:
        ma = ma_var.get().strip()
        if not ma:
            return rows
        key = clean_key(ma)
        out: list[dict] = []
        for r in rows:
            prod_key = clean_key(r.get("production_no", ""))
            if key and (key == prod_key or key in prod_key):
                out.append(r)
        return out

    def render_check_main_table(self, result_df: pd.DataFrame | None = None) -> None:
        self.current_view_mode = "summary"
        base = result_df if result_df is not None else self.last_result_df
        if base is None or base.empty:
            self.check_page = 1
            self._set_check_view_rows([])
            return
        view_df = self._filter_result_df_by_o(base)
        summary_rows = self._summary_rows_from_df(view_df, self.run_case_meta_by_dg)
        summary_rows = self._filter_summary_rows_by_production(summary_rows, self.check_filter_ma_var)
        summary_rows = self._filter_summary_rows_by_status(summary_rows, self.check_filter_status_var)
        self.check_page = 1
        out: list[dict] = []
        for r in summary_rows:
            out.append(
                {
                    "values": (r["dg"], r["production_no"], r["ship_date"], r["ket_qua"], r["ket_luan"], r["ghi_chu"]),
                    "tag": r["tag"],
                }
            )
        self._set_check_view_rows(out)

    def apply_history_item_filters(self) -> None:
        if self.history_last_df is None or self.history_last_df.empty:
            messagebox.showinfo("Thông báo", "Chọn một run và tải chi tiết trước.")
            return
        self._render_history_items_from_df(self.history_last_df, self.history_current_run_type or "one")

    def choose_order_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Chọn file ORDER LIST",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if path:
            self.order_file_var.set(path)
            self._save_last_paths()
            self._refresh_customer_filter_options(path)

    def _refresh_customer_filter_options(self, order_file: str) -> None:
        options = [CUSTOMER_FILTER_EMG, CUSTOMER_FILTER_NON_EMG, CUSTOMER_FILTER_ALL]
        try:
            if order_file and Path(order_file).exists():
                df = pd.read_excel(order_file, sheet_name=0, header=0, usecols=[5])
                names = sorted(
                    {
                        clean_text(v)
                        for v in df.iloc[:, 0].tolist()
                        if clean_text(v) and clean_key(v) != "emg"
                    }
                )
                options.extend(names)
        except Exception:
            # Giữ fallback mặc định nếu không đọc được file Order.
            pass
        self.customer_filter_values = options
        self.customer_filter_combo.configure(values=self.customer_filter_values)
        if self.customer_filter_var.get() not in self.customer_filter_values:
            self.customer_filter_var.set(CUSTOMER_FILTER_EMG)

    def _history_date_from_iso(self) -> str:
        w = getattr(self, "history_date_from", None)
        if w is not None and DateEntry is not None:
            try:
                return w.get_date().strftime("%Y-%m-%d")
            except (tk.TclError, ValueError, AttributeError):
                return ""
        return self.date_from_var.get().strip()

    def _history_date_to_iso(self) -> str:
        w = getattr(self, "history_date_to", None)
        if w is not None and DateEntry is not None:
            try:
                return w.get_date().strftime("%Y-%m-%d")
            except (tk.TclError, ValueError, AttributeError):
                return ""
        return self.date_to_var.get().strip()

    def _build_history_tab(self) -> None:
        self.history_date_from = None
        self.history_date_to = None
        filter_wrap = ttk.Frame(self.tab_history, padding=10)
        filter_wrap.pack(fill="x")
        ttk.Label(filter_wrap, text="Từ ngày:").pack(side="left")
        if DateEntry is not None:
            self.history_date_from = DateEntry(
                filter_wrap,
                width=11,
                date_pattern="y-mm-dd",
                year_from=2018,
                year_to=2038,
                firstweekday="monday",
            )
            self.history_date_from.pack(side="left", padx=(6, 10))
            self.history_date_from.bind("<<DateEntrySelected>>", lambda _e: self.refresh_history_runs())
        else:
            ttk.Label(
                filter_wrap,
                text="(cài: pip install tkcalendar)",
                font=("Segoe UI", 8),
                foreground="#666666",
            ).pack(side="left", padx=(0, 4))
            ttk.Entry(filter_wrap, textvariable=self.date_from_var, width=14).pack(side="left", padx=(6, 10))
        ttk.Label(filter_wrap, text="Đến ngày:").pack(side="left")
        if DateEntry is not None:
            self.history_date_to = DateEntry(
                filter_wrap,
                width=11,
                date_pattern="y-mm-dd",
                year_from=2018,
                year_to=2038,
                firstweekday="monday",
            )
            self.history_date_to.pack(side="left", padx=(6, 10))
            self.history_date_to.bind("<<DateEntrySelected>>", lambda _e: self.refresh_history_runs())
        else:
            ttk.Entry(filter_wrap, textvariable=self.date_to_var, width=14).pack(side="left", padx=(6, 10))
        ttk.Button(filter_wrap, text="Hôm nay", command=self._set_default_dates).pack(side="left", padx=(6, 0))
        ttk.Button(filter_wrap, text="Ghim / Bỏ ghim", command=self.toggle_pin_history_run).pack(side="left", padx=(10, 0))
        ttk.Label(filter_wrap, text="Lọc tên run:").pack(side="left", padx=(18, 4))
        ttk.Entry(filter_wrap, textvariable=self.hist_filter_run_label_var, width=24).pack(side="left", padx=(0, 8))

        hist_filter_row = ttk.Frame(self.tab_history, padding=(10, 0, 10, 4))
        hist_filter_row.pack(fill="x")
        ttk.Label(hist_filter_row, text="Lọc số O (DG):").pack(side="left")
        ttk.Entry(hist_filter_row, textvariable=self.hist_filter_o_var, width=22).pack(side="left", padx=(6, 14))
        ttk.Label(hist_filter_row, text="Lọc mã:").pack(side="left")
        ttk.Entry(hist_filter_row, textvariable=self.hist_filter_ma_var, width=18).pack(side="left", padx=(6, 14))
        ttk.Label(hist_filter_row, text="Trạng thái:").pack(side="left")
        ttk.Combobox(
            hist_filter_row,
            textvariable=self.hist_filter_status_var,
            values=CHECK_STATUS_FILTER_VALUES,
            width=22,
            state="readonly",
        ).pack(side="left", padx=(6, 10))
        ttk.Button(hist_filter_row, text="Thiết lập Quy tắc", command=self.open_rule_setup_from_history_tab).pack(
            side="left", padx=(8, 0)
        )

        run_cols = (
            "run_id",
            "created_at",
            "run_label",
            "run_type",
            "trace_ol",
            "trace_shipped",
            "trace_bk_a6",
            "total",
            "lech",
            "manual_check",
            "ghim",
        )
        run_wrap = ttk.Frame(self.tab_history, padding=(10, 0, 10, 6))
        run_wrap.pack(fill="both", expand=True)
        self.history_runs_tree = ttk.Treeview(run_wrap, columns=run_cols, show="headings", height=10)
        run_headings = {
            "run_id": "run_id",
            "created_at": "created_at",
            "run_label": "Tên run / ghi chú",
            "run_type": "run_type",
            "trace_ol": "OL Name",
            "trace_shipped": "Shipped List",
            "trace_bk_a6": "Bảng kê",
            "total": "total",
            "lech": "lech",
            "manual_check": "Check",
            "ghim": "Ghim",
        }
        for col, width, stretch in [
            ("run_id", 56, False),
            ("created_at", 138, False),
            ("run_label", 160, False),
            ("run_type", 72, False),
            ("trace_ol", 200, False),
            ("trace_shipped", 150, False),
            ("trace_bk_a6", 360, True),
            ("total", 64, False),
            ("lech", 56, False),
            ("manual_check", 56, False),
            ("ghim", 44, False),
        ]:
            self.history_runs_tree.heading(col, text=run_headings[col])
            self.history_runs_tree.column(col, width=width, anchor="center", stretch=stretch)
        self.history_runs_tree.pack(side="left", fill="both", expand=True)
        run_scroll = ttk.Scrollbar(run_wrap, orient="vertical", command=self.history_runs_tree.yview)
        run_scroll.pack(side="right", fill="y")
        self.history_runs_tree.configure(yscrollcommand=run_scroll.set)
        self.history_runs_tree.tag_configure("pinned", background="#fff3e0")
        self.history_runs_tree.bind("<Double-1>", self.load_history_run_items)

        detail_cols = ("dg_case_no", "production_no", "ship_date", "ket_qua", "ket_luan", "ghi_chu")
        hist_head = {
            "dg_case_no": "DG Case",
            "production_no": "Production No.",
            "ship_date": "Ship date",
            "ket_qua": "Kết quả check",
            "ket_luan": "Kết luận",
            "ghi_chu": "Ghi chú",
        }
        detail_wrap = ttk.Frame(self.tab_history, padding=(10, 0, 10, 10))
        detail_wrap.pack(fill="both", expand=True)
        self.history_items_tree = ttk.Treeview(detail_wrap, columns=detail_cols, show="headings", height=12)
        for col, width in [
            ("dg_case_no", 140),
            ("production_no", 120),
            ("ship_date", 100),
            ("ket_qua", 120),
            ("ket_luan", 160),
            ("ghi_chu", 280),
        ]:
            self.history_items_tree.heading(col, text=hist_head[col])
            self.history_items_tree.column(col, width=width, anchor="center")
        self.history_items_tree.pack(side="left", fill="both", expand=True)
        detail_scroll = ttk.Scrollbar(detail_wrap, orient="vertical", command=self.history_items_tree.yview)
        detail_scroll.pack(side="right", fill="y")
        self.history_items_tree.configure(yscrollcommand=detail_scroll.set)
        self.history_items_tree.tag_configure("c_ok", background="#a5d6a7")
        self.history_items_tree.tag_configure("m_ok", background="#b8d4f0")
        self.history_items_tree.tag_configure("m_bad", background="#fff59d")
        self.history_items_tree.tag_configure("rule_bad", background="#ffcc80")
        self.history_items_tree.tag_configure("c_bad", background="#ef9a9a")
        self.history_items_tree.bind("<Double-1>", self.on_history_items_double_click)

    def _set_default_dates(self) -> None:
        today_d = datetime.now().date()
        today_s = today_d.strftime("%Y-%m-%d")
        self.date_from_var.set(today_s)
        self.date_to_var.set(today_s)
        if getattr(self, "history_date_from", None) is not None:
            try:
                self.history_date_from.set_date(today_d)
            except tk.TclError:
                pass
        if getattr(self, "history_date_to", None) is not None:
            try:
                self.history_date_to.set_date(today_d)
            except tk.TclError:
                pass
        if hasattr(self, "history_runs_tree"):
            self.refresh_history_runs()

    def _start_clock(self) -> None:
        self.clock_var.set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.root.after(1000, self._start_clock)

    def toggle_pin_history_run(self) -> None:
        if not hasattr(self, "history_runs_tree"):
            return
        sel = self.history_runs_tree.selection()
        if not sel:
            messagebox.showinfo("Ghim", "Chọn một run trong danh sách trên.")
            return
        vals = self.history_runs_tree.item(sel[0], "values")
        if not vals:
            return
        try:
            rid = int(vals[0])
        except (TypeError, ValueError):
            return
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("PRAGMA foreign_keys = ON")
        ex = cur.execute("SELECT 1 FROM pinned_runs WHERE run_id = ?", (rid,)).fetchone()
        if ex:
            cur.execute("DELETE FROM pinned_runs WHERE run_id = ?", (rid,))
            msg = "Đã bỏ ghim run này."
        else:
            cur.execute("INSERT OR IGNORE INTO pinned_runs (run_id) VALUES (?)", (rid,))
            msg = "Đã ghim run — run luôn hiện ở đầu danh sách, không bị lọc ngày/tên loại trừ."
        conn.commit()
        conn.close()
        messagebox.showinfo("Ghim", msg)
        self.refresh_history_runs()

    def refresh_history_runs(self) -> None:
        if not hasattr(self, "history_runs_tree"):
            return
        for item in self.history_runs_tree.get_children():
            self.history_runs_tree.delete(item)
        for item in self.history_items_tree.get_children():
            self.history_items_tree.delete(item)

        date_from = self._history_date_from_iso()
        date_to = self._history_date_to_iso()
        where = []
        params: list[str] = []
        run_name_sub = self.hist_filter_run_label_var.get().strip()
        if run_name_sub:
            where.append("instr(lower(COALESCE(r.run_label, '')), lower(?)) > 0")
            params.append(run_name_sub)
        else:
            if date_from:
                where.append("date(r.created_at) >= date(?)")
                params.append(date_from)
            if date_to:
                where.append("date(r.created_at) <= date(?)")
                params.append(date_to)
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""

        run_agg_sql = """
            SELECT
                r.id,
                r.created_at,
                COALESCE(r.run_label, ''),
                r.run_type,
                COALESCE(r.trace_orderlist_filename, ''),
                COALESCE(r.trace_shipped_file_mtime, ''),
                COALESCE(r.trace_bang_ke_a6, ''),
                COUNT(i.id) AS total_checks,
                SUM(CASE WHEN i.status = 'Lệch' THEN 1 ELSE 0 END) AS bad_checks,
                (SELECT COUNT(*) FROM run_case_meta m WHERE m.run_id = r.id) AS manual_case_checks
            FROM runs r
            LEFT JOIN run_items i ON i.run_id = r.id
        """

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("PRAGMA foreign_keys = ON")
        cur.execute("SELECT run_id FROM pinned_runs")
        pinned_ids = {int(x[0]) for x in cur.fetchall()}

        rows_filtered = list(
            cur.execute(
                f"{run_agg_sql} {where_sql} GROUP BY r.id ORDER BY r.id DESC",
                params,
            ).fetchall()
        )
        ids_filtered = {int(r[0]) for r in rows_filtered}
        extra_ids = sorted(pinned_ids - ids_filtered, reverse=True)
        extra_rows: list[tuple] = []
        for rid in extra_ids:
            row = cur.execute(
                f"{run_agg_sql} WHERE r.id = ? GROUP BY r.id",
                (rid,),
            ).fetchone()
            if row:
                extra_rows.append(row)
        conn.close()

        for row in extra_rows:
            self.history_runs_tree.insert(
                "",
                "end",
                values=(*row, "★"),
                tags=("pinned",),
            )
        for row in rows_filtered:
            pin_mark = "★" if int(row[0]) in pinned_ids else ""
            tags = ("pinned",) if pin_mark else ()
            self.history_runs_tree.insert("", "end", values=(*row, pin_mark), tags=tags)

    def load_history_run_items(self, _event: tk.Event | None = None) -> None:
        selected = self.history_runs_tree.selection()
        if not selected:
            return
        run_vals = self.history_runs_tree.item(selected[0], "values")
        if not run_vals:
            return
        run_id = int(run_vals[0])
        run_type = str(run_vals[3])
        self.history_current_run_id = run_id
        self.history_current_run_type = run_type
        for item in self.history_items_tree.get_children():
            self.history_items_tree.delete(item)

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        bk_row = cur.execute("SELECT bang_ke_file FROM runs WHERE id = ?", (run_id,)).fetchone()
        self.history_selected_bang_ke_file = clean_text(bk_row[0]) if bk_row else ""
        rows = cur.execute(
            """
            SELECT dg_case_no, field_name, order_value, COALESCE(shipped_value, ''), bang_ke_value, status,
                   COALESCE(adjust_reason, ''),
                   COALESCE(production_no, ''), COALESCE(ship_date_display, '')
            FROM run_items
            WHERE run_id = ?
            ORDER BY dg_case_no, field_name
            """,
            (run_id,),
        ).fetchall()
        meta_rows = cur.execute(
            "SELECT dg_case_no, COALESCE(user_conclusion, ''), COALESCE(case_note, '') FROM run_case_meta WHERE run_id = ?",
            (run_id,),
        ).fetchall()
        conn.close()
        self.history_case_meta_by_dg = {
            clean_text(r[0]): {"user_conclusion": clean_text(r[1]), "case_note": clean_text(r[2])} for r in meta_rows
        }
        self.history_last_df = pd.DataFrame(
            rows,
            columns=[
                "dg_case_no",
                "field_name",
                "order_value",
                "shipped_value",
                "bang_ke_value",
                "status_core",
                "adjust_reason",
                "production_no",
                "ship_date_display",
            ],
        )
        if self.history_last_df.empty:
            self.history_case_meta_by_dg = {}
            return
        self._render_history_items_from_df(self.history_last_df, run_type)

    def _render_history_items_from_df(self, df: pd.DataFrame, _run_type: str) -> None:
        for item in self.history_items_tree.get_children():
            self.history_items_tree.delete(item)
        if df.empty:
            return
        view_df = self._filter_hist_df_by_o(df)
        summary_rows = self._summary_rows_from_df(view_df, self.history_case_meta_by_dg)
        summary_rows = self._filter_summary_rows_by_production(summary_rows, self.hist_filter_ma_var)
        summary_rows = self._filter_summary_rows_by_status(summary_rows, self.hist_filter_status_var)
        for r in summary_rows:
            self.history_items_tree.insert(
                "",
                "end",
                values=(r["dg"], r["production_no"], r["ship_date"], r["ket_qua"], r["ket_luan"], r["ghi_chu"]),
                tags=(r["tag"],),
            )

    def on_history_items_double_click(self, _event: tk.Event | None = None) -> None:
        if self.history_last_df is None or self.history_last_df.empty:
            return
        selected = self.history_items_tree.selection()
        if not selected:
            return
        vals = self.history_items_tree.item(selected[0], "values")
        if not vals:
            return
        dg_case = str(vals[0]).strip()
        if not dg_case:
            return
        detail_df = self.history_last_df[self.history_last_df["dg_case_no"] == dg_case]
        if detail_df.empty:
            return
        self.open_detail_window(
            dg_case,
            detail_df,
            self.history_selected_bang_ke_file,
            persist_run_id=self.history_current_run_id,
            meta_by_dg=self.history_case_meta_by_dg,
        )

    def choose_bang_ke_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Chọn file Bảng Kê định mức",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if path:
            self.bang_ke_file_var.set(path)
            self._save_last_paths()

    def choose_shipped_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Chọn file SHIPPED LIST",
            filetypes=[("Excel files", "*.xlsx *.xls")],
        )
        if path:
            self.shipped_file_var.set(path)
            self._save_last_paths()

    def _load_last_paths(self) -> None:
        self.order_file_var.set(self.config.get("order_file", ""))
        self.shipped_file_var.set(self.config.get("shipped_file", ""))
        self.bang_ke_file_var.set(self.config.get("bang_ke_file", ""))
        self.customer_filter_var.set(CUSTOMER_FILTER_EMG)
        self._refresh_customer_filter_options(self.order_file_var.get().strip())

    def _save_last_paths(self) -> None:
        self.config["order_file"] = self.order_file_var.get().strip()
        self.config["shipped_file"] = self.shipped_file_var.get().strip()
        self.config["bang_ke_file"] = self.bang_ke_file_var.get().strip()
        save_config(self.config)

    def run_all(self) -> None:
        self._run(run_type="all", target_dg=None)

    def run_one(self) -> None:
        dg_case = self.dg_case_var.get().strip()
        if not dg_case:
            messagebox.showwarning("Thiếu DG Case", "Nhập DG Case No để chạy Run One.")
            return
        self._run(run_type="one", target_dg=dg_case)

    def _run(self, run_type: str, target_dg: str | None) -> None:
        order_file = self.order_file_var.get().strip()
        shipped_file = self.shipped_file_var.get().strip()
        bang_ke_file = self.bang_ke_file_var.get().strip()
        if not order_file or not shipped_file or not bang_ke_file:
            messagebox.showwarning("Thiếu file", "Hãy chọn đủ file ORDER LIST, SHIPPED LIST và Bảng Kê.")
            return
        if not Path(order_file).exists() or not Path(shipped_file).exists() or not Path(bang_ke_file).exists():
            messagebox.showerror("Sai đường dẫn", "Một trong ba file không tồn tại.")
            return

        run_label = ""
        if run_type == "all":
            run_label = simpledialog.askstring(
                "Tên run (Run All)",
                "Nhập tên hoặc ghi chú (hiện trong Lịch sử để dễ tìm).\n"
                "Để trống nếu không cần — bấm OK để chạy.\n"
                "Bấm Cancel để hủy không chạy.",
                parent=self.root,
            )
            if run_label is None:
                self.status_var.set("Đã hủy Run All.")
                return
            run_label = clean_text(run_label)

        self.status_var.set("Đang xử lý dữ liệu...")
        self.root.update_idletasks()
        try:
            self.last_run_type = run_type
            self.last_target_dg = target_dg
            self.last_run_customer_filter = self.customer_filter_var.get().strip() or CUSTOMER_FILTER_EMG
            result_df = self.compare_files(
                order_file,
                shipped_file,
                bang_ke_file,
                target_dg,
                self.customer_filter_var.get().strip(),
            )
            if result_df.empty:
                self.current_run_id = None
                self.last_result_df = result_df
                self.run_case_meta_by_dg = {}
                self.render_check_main_table(result_df)
                self.status_var.set(
                    "Không có dữ liệu khớp điều kiện (khách đã chọn + DG + Ship date tương lai + Bảng kê)."
                )
                return
            run_id = self.save_run(
                run_type,
                target_dg,
                order_file,
                shipped_file,
                bang_ke_file,
                result_df,
                run_label=run_label,
            )
            self.current_run_id = run_id
            self.last_result_df = result_df
            self.run_case_meta_by_dg = {}
            self.render_check_main_table(result_df)
            ok_count = int((result_df["status_core"] == "Đúng").sum())
            bad_count = int((result_df["status_core"] == "Lệch").sum())
            msg = (
                f"Hoàn tất. Run ID {run_id} | Tổng dòng check: {len(result_df)} | Đúng: {ok_count} | Lệch: {bad_count}"
                " | Double-click DG → chi tiết tiêu chí; xác nhận + ghi chú ở panel dưới cùng; «Tải phụ liệu» cho NPL."
            )
            self.status_var.set(msg)
            self.refresh_history_runs()
        except Exception as exc:
            messagebox.showerror("Lỗi", str(exc))
            self.status_var.set("Có lỗi khi xử lý.")
        finally:
            self._run_progress_done()

    def _refresh_check_tab_current_data(self, reason: str) -> None:
        if self.last_result_df is None:
            return
        order_file = self.order_file_var.get().strip()
        shipped_file = self.shipped_file_var.get().strip()
        bang_ke_file = self.bang_ke_file_var.get().strip()
        if not order_file or not shipped_file or not bang_ke_file:
            return
        if not Path(order_file).exists() or not Path(shipped_file).exists() or not Path(bang_ke_file).exists():
            return
        try:
            self.status_var.set("Đang cập nhật tab Check theo thiết lập mới...")
            self.root.update_idletasks()
            refreshed_df = self.compare_files(
                order_file,
                shipped_file,
                bang_ke_file,
                self.last_target_dg if self.last_run_type == "one" else None,
                self.last_run_customer_filter,
            )
            self.last_result_df = refreshed_df
            self.render_check_main_table(refreshed_df)
            self.status_var.set(
                f"Đã cập nhật tab Check sau khi lưu {reason}. Dòng check: {len(refreshed_df)}"
            )
        except Exception as exc:
            self.status_var.set(f"Đã lưu {reason}, nhưng chưa refresh được tab Check: {exc}")

    def compare_files(
        self,
        order_file: str,
        shipped_file: str,
        bang_ke_file: str,
        target_dg: str | None,
        run_all_customer_filter: str,
    ) -> pd.DataFrame:
        self._run_progress_io_begin()
        try:
            self.status_var.set("Đang đọc ORDER LIST…")
            self.root.update()
            order_df = pd.read_excel(order_file, sheet_name=0, header=0)
            self.status_var.set("Đang đọc SHIPPED LIST…")
            self.root.update()
            shipped_df = pd.read_excel(shipped_file, sheet_name=0, header=0)
            self.status_var.set("Đang đọc Bảng kê (tìm dòng tiêu đề)…")
            self.root.update()
            header_row = find_bang_ke_header_row(bang_ke_file)
            self.status_var.set("Đang đọc Bảng kê + chuẩn bị tra cứu…")
            self.root.update()
            bang_ke_df = pd.read_excel(bang_ke_file, sheet_name=0, header=header_row)
            bang_ke_df = annotate_bang_ke_for_fast_lookup(bang_ke_df)
        finally:
            self._run_progress_io_end()

        self._store_bang_ke_in_path_cache(bang_ke_file, bang_ke_df)

        # ORDER LIST + SHIPPED LIST: luôn giữ DG có dữ liệu; Run All mới áp filter khách.
        def preprocess_order_like(df: pd.DataFrame) -> pd.DataFrame:
            out = df.copy()
            out["customer_name"] = out.iloc[:, 5].map(clean_text)
            out["customer_key"] = out.iloc[:, 5].map(clean_key)
            out["dg_case"] = out.iloc[:, 2].map(clean_text)
            return out[out["dg_case"] != ""]

        order_df = preprocess_order_like(order_df)
        shipped_df = preprocess_order_like(shipped_df)
        if target_dg:
            target_key = clean_key(target_dg)
            order_df = order_df[order_df["dg_case"].apply(clean_key) == target_key]
            shipped_df = shipped_df[shipped_df["dg_case"].apply(clean_key) == target_key]
        else:
            selected = clean_text(run_all_customer_filter) or CUSTOMER_FILTER_EMG
            selected_key = clean_key(selected)
            if selected == CUSTOMER_FILTER_EMG:
                order_df = order_df[order_df["customer_key"] == "emg"]
                shipped_df = shipped_df[shipped_df["customer_key"] == "emg"]
            elif selected == CUSTOMER_FILTER_NON_EMG:
                order_df = order_df[(order_df["customer_key"] != "") & (order_df["customer_key"] != "emg")]
                shipped_df = shipped_df[(shipped_df["customer_key"] != "") & (shipped_df["customer_key"] != "emg")]
            elif selected == CUSTOMER_FILTER_ALL:
                pass
            else:
                order_df = order_df[order_df["customer_key"] == selected_key]
                shipped_df = shipped_df[shipped_df["customer_key"] == selected_key]

        if order_df.empty and shipped_df.empty:
            return pd.DataFrame(
                columns=[
                    "dg_case_no",
                    "field_name",
                    "order_value",
                    "shipped_value",
                    "bang_ke_value",
                    "auto_status",
                    "status_core",
                    "adjust_reason",
                    "production_no",
                    "ship_date_display",
                    "_sort_days",
                ]
            )

        if target_dg:
            # Run One: cho phép chạy nếu chỉ có trong SHIPPED.
            all_dg_keys = set(order_df["dg_case"].dropna().astype(str).tolist()) | set(
                shipped_df["dg_case"].dropna().astype(str).tolist()
            )
        else:
            # Run All: chỉ chạy các mã thực sự có trong ORDER LIST.
            all_dg_keys = set(order_df["dg_case"].dropna().astype(str).tolist())
        records: list[dict] = []

        def append_error_record(
            dg_case_value: object,
            group_df: pd.DataFrame,
            stage: str,
            exc: Exception,
            compare_target: str = "",
            order_g: pd.DataFrame | None = None,
            shipped_g: pd.DataFrame | None = None,
        ) -> None:
            dg_text = clean_text(dg_case_value)
            if not dg_text:
                dg_text = "(DG trống)"
            order_rows = ", ".join(str(int(idx) + 2) for idx in group_df.index.tolist()[:8])
            if len(group_df.index) > 8:
                order_rows += ", ..."
            if not order_rows:
                order_rows = "(không xác định)"
            detail = f"{type(exc).__name__}: {clean_text(str(exc))}"
            field = f"LỖI XỬ LÝ ({stage})"
            if clean_text(compare_target):
                field = f"{field} | So: {clean_text(compare_target)}"
            og = order_g if order_g is not None else group_df
            sg = shipped_g if shipped_g is not None else group_df
            prod_err = ol_production_no_from_groups(og, sg)
            ship_disp_err, sort_err = ol_ship_date_display_and_sort_days(group_df)
            records.append(
                {
                    "dg_case_no": dg_text,
                    "field_name": field,
                    "order_value": f"ORDER LIST dòng: {order_rows}",
                    "shipped_value": "",
                    "bang_ke_value": detail,
                    "auto_status": "Lỗi",
                    "status_core": "Lệch",
                    "adjust_reason": f"Lỗi tại bước {stage}" + (f" khi so {compare_target}." if clean_text(compare_target) else "."),
                    "production_no": prod_err,
                    "ship_date_display": ship_disp_err,
                    "_sort_days": sort_err,
                }
            )

        sorted_dg_list = sorted(all_dg_keys)
        use_progress = target_dg is None and len(sorted_dg_list) > 0
        if use_progress:
            self._run_progress_start(len(sorted_dg_list))

        order_by_dg: dict[str, pd.DataFrame] = {
            str(k): v for k, v in order_df.groupby("dg_case", dropna=False, sort=False)
        }
        shipped_by_dg: dict[str, pd.DataFrame] = {
            str(k): v for k, v in shipped_df.groupby("dg_case", dropna=False, sort=False)
        }
        empty_order = order_df.iloc[0:0].copy()
        empty_shipped = shipped_df.iloc[0:0].copy()
        has_bk_aux = "_dg_key" in bang_ke_df.columns
        if has_bk_aux:
            bk_by_dg: dict[str, pd.DataFrame] = {
                str(k): g
                for k, g in bang_ke_df.groupby("_dg_key", dropna=False, sort=False)
                if clean_text(str(k))
            }
        else:
            bk_by_dg = {}
        empty_bk = bang_ke_df.iloc[0:0].copy()

        for dg_i, dg_case in enumerate(sorted_dg_list):
            if use_progress and self._run_progress_total > 0 and (
                dg_i % 2 == 0 or dg_i == len(sorted_dg_list) - 1
            ):
                self._run_progress_update(dg_i + 1)
            dg_key_str = str(dg_case)
            order_group = order_by_dg.get(dg_key_str, empty_order)
            shipped_group = shipped_by_dg.get(dg_key_str, empty_shipped)
            current_compare_target = ""
            try:
                base_group_for_ship_date = order_group if not order_group.empty else shipped_group
                if base_group_for_ship_date.empty:
                    continue
                if not group_has_future_ship_date(base_group_for_ship_date):
                    continue
            except Exception as exc:
                append_error_record(
                    dg_case,
                    order_group if not order_group.empty else shipped_group,
                    "lọc ship date",
                    exc,
                    "",
                    order_group,
                    shipped_group,
                )
                continue

            try:
                current_compare_target = "đọc dữ liệu ORDER LIST và SHIPPED LIST theo DG"
                order_metrics = extract_order_like_metrics(order_group)
                shipped_metrics = extract_order_like_metrics(shipped_group)
                ol_missing = order_group.empty
                sum_qty_total = float(order_metrics["qty_total"]) + float(shipped_metrics["qty_total"])
                sum_carton_qty = float(order_metrics["carton_qty"]) + float(shipped_metrics["carton_qty"])
                prod_no = ol_production_no_from_groups(order_group, shipped_group)
                ship_date_display, sort_days = ol_ship_date_display_and_sort_days(base_group_for_ship_date)

                current_compare_target = "match DG giữa Order List và Bảng kê (cột đầu Bảng kê)"
                lookup_bk = clean_key(dg_case)
                if has_bk_aux:
                    bang_ke_match = bk_by_dg.get(lookup_bk, empty_bk)
                else:
                    bang_ke_match = bang_ke_df[bang_ke_df.iloc[:, 0].map(clean_key) == lookup_bk]
                if bang_ke_match.empty:
                    records.append(
                        {
                            "dg_case_no": clean_text(dg_case),
                            "field_name": "LỖI XỬ LÝ (đối chiếu dữ liệu) | So: match DG với Bảng kê",
                            "order_value": (
                                "ORDER LIST dòng: "
                                + ", ".join(str(int(r) + 2) for r in order_group.index.tolist()[:8])
                                + " | SHIPPED LIST dòng: "
                                + ", ".join(str(int(r) + 2) for r in shipped_group.index.tolist()[:8])
                            ),
                            "shipped_value": "",
                            "bang_ke_value": "Không có dữ liệu trong bảng kê",
                            "auto_status": "Lỗi",
                            "status_core": "Lệch",
                            "adjust_reason": "Không có dữ liệu trong bảng kê",
                            "production_no": prod_no,
                            "ship_date_display": ship_date_display,
                            "_sort_days": sort_days,
                        }
                    )
                    continue
                current_compare_target = "đọc Bảng kê: Đơn hàng / Logo / Mã SP / Tên SP"
                bk_note = mode_value(bang_ke_match.iloc[:, 8]) if not bang_ke_match.empty else ""
                bk_logo = normalize_logo(mode_value(bang_ke_match.iloc[:, 7])) if not bang_ke_match.empty else ""
                bk_logo_type = logo_type_from_bang_ke_npl(bang_ke_match)
                bk_ma_sp = mode_value(bang_ke_match.iloc[:, 3]) if not bang_ke_match.empty else ""
                bk_ten_sp = mode_value(bang_ke_match.iloc[:, 4]) if not bang_ke_match.empty else ""

                # Carton (thùng) và Pallet: tách 2 luật, luôn có 2 nhóm dòng so riêng.
                pallet_rows = pd.DataFrame()
                carton_only_rows = pd.DataFrame()
                if not bang_ke_match.empty:
                    current_compare_target = "lọc dòng pallet theo NPL (Bảng kê cột 9)"
                    if has_bk_aux and "_pallet948" in bang_ke_match.columns:
                        pallet_rows = bang_ke_match[bang_ke_match["_pallet948"]]
                        current_compare_target = "lọc dòng carton theo mô tả NPL (Bảng kê cột 10/11)"
                        carton_only_rows = bang_ke_match[
                            bang_ke_match["_carton_row"] & ~bang_ke_match["_pallet948"]
                        ]
                    else:
                        pallet_mask = bang_ke_match.iloc[:, 9].map(lambda v: "948pallet" in clean_key(v))
                        pallet_rows = bang_ke_match[pallet_mask]
                        current_compare_target = "lọc dòng carton theo mô tả NPL (Bảng kê cột 10/11)"
                        carton_mask = bang_ke_match.apply(
                            lambda row: (
                                ("cartonbox" in clean_key(row.iloc[10]))
                                or ("cartonbox" in clean_key(row.iloc[11]))
                                or ("carton" in clean_key(row.iloc[10]))
                                or ("carton" in clean_key(row.iloc[11]))
                            ),
                            axis=1,
                        )
                        carton_only_rows = bang_ke_match[carton_mask & ~pallet_mask]

                current_compare_target = "đọc số lượng carton/pallet từ Bảng kê (cột 6)"
                bk_qty_carton = mode_value(carton_only_rows.iloc[:, 6]) if not carton_only_rows.empty else ""
                bk_qty_pallet = mode_value(pallet_rows.iloc[:, 6]) if not pallet_rows.empty else ""

                current_compare_target = "đọc số thùng carton từ Bảng kê (cột 15)"
                carton_qty_bk = mode_value(carton_only_rows.iloc[:, 15]) if not carton_only_rows.empty else ""
                current_compare_target = "đọc size carton từ NPL Bảng kê (cột 9)"
                carton_npl_bk = mode_value(carton_only_rows.iloc[:, 9]) if not carton_only_rows.empty else ""
                npl_x, npl_y, npl_z = parse_npl_950_code(carton_npl_bk)
                bang_ke_size_xyz_carton = ".".join(v for v in [npl_x, npl_y, npl_z] if clean_text(v))

                current_compare_target = "cộng số thùng pallet từ Bảng kê (cột 15)"
                pallet_nums = [v for v in pallet_rows.iloc[:, 15].apply(to_number) if v is not None]
                pallet_qty_bk = format_number(sum(pallet_nums)) if pallet_nums else ""
                has_pallet = not pallet_rows.empty
                fabric_rows = pd.DataFrame()
                if not bang_ke_match.empty:
                    current_compare_target = "lọc dòng vải theo mô tả Bảng kê (cột 11)"
                    if has_bk_aux and "_fabric" in bang_ke_match.columns:
                        fabric_rows = bang_ke_match[bang_ke_match["_fabric"]].copy()
                    else:
                        fabric_rows = bang_ke_match[
                            bang_ke_match.iloc[:, 11].map(lambda v: "vai" in clean_key(v))
                        ].copy()
                fabric_color_code = ""
                fabric_color_name = ""
                fabric_npl_full = ""
                fabric_npl_desc = ""
                if not fabric_rows.empty:
                    current_compare_target = "ưu tiên dòng có cụm 'vải chính', nếu không có thì lấy qty lớn nhất"
                    fabric_rows["qty_p"] = fabric_rows.iloc[:, 15].apply(to_number).fillna(0)
                    fabric_main_mask = fabric_rows.apply(
                        lambda row: ("vaichinh" in clean_key(row.iloc[10])) or ("vaichinh" in clean_key(row.iloc[11])),
                        axis=1,
                    )
                    fabric_candidates = fabric_rows[fabric_main_mask].copy()
                    if fabric_candidates.empty:
                        fabric_candidates = fabric_rows
                    fabric_top = fabric_candidates.sort_values("qty_p", ascending=False).iloc[0]
                    fabric_color_code = parse_npl_color(fabric_top.iloc[9])
                    fabric_color_name = color_name_from_code(fabric_color_code, self.color_pairs)
                    fabric_npl_full = clean_text(fabric_top.iloc[9])
                    fabric_npl_desc = clean_text(fabric_top.iloc[10])

                color_display = f"{fabric_npl_full} | {fabric_npl_desc} ({fabric_color_name})".strip(" |")
                if not fabric_npl_full and not fabric_npl_desc:
                    color_display = ""

                order_size_xyz = ".".join(
                    v
                    for v in [
                        format_size_part_rounded(order_metrics["color_s"]),
                        format_size_part_rounded(order_metrics["size_t"]),
                        format_size_part_rounded(order_metrics["size_u"]),
                    ]
                    if clean_text(v)
                )
                shipped_size_xyz = ".".join(
                    v
                    for v in [
                        format_size_part_rounded(shipped_metrics["color_s"]),
                        format_size_part_rounded(shipped_metrics["size_t"]),
                        format_size_part_rounded(shipped_metrics["size_u"]),
                    ]
                    if clean_text(v)
                )

                checks = [
                    (
                        "Đơn hàng",
                        clean_text(order_metrics["order_no"]),
                        clean_text(shipped_metrics["order_no"]),
                        clean_text(bk_note),
                    ),
                    (
                        "Số lượng đơn hàng",
                        format_number(float(order_metrics["qty_total"])),
                        format_number(float(shipped_metrics["qty_total"])),
                        clean_text(bk_qty_pallet if has_pallet else bk_qty_carton),
                    ),
                    (
                        "Logo",
                        clean_text(order_metrics["logo"]),
                        clean_text(shipped_metrics["logo"]),
                        clean_text(bk_logo),
                    ),
                    (
                        "Loại logo",
                        logo_type_to_label(order_metrics["logo_type"]),
                        logo_type_to_label(shipped_metrics["logo_type"]),
                        logo_type_to_label(bk_logo_type),
                    ),
                    (
                        "Mã sản phẩm",
                        clean_text(order_metrics["ma_sp"]),
                        clean_text(shipped_metrics["ma_sp"]),
                        clean_text(bk_ma_sp),
                    ),
                    (
                        "Tên sản phẩm",
                        clean_text(order_metrics["ten_sp"]),
                        clean_text(shipped_metrics["ten_sp"]),
                        clean_text(bk_ten_sp),
                    ),
                    (
                        "Màu sắc",
                        clean_text(order_metrics["color_k"]),
                        clean_text(shipped_metrics["color_k"]),
                        clean_text(color_display),
                    ),
                    (
                        "Số thùng (carton)",
                        format_number(float(order_metrics["carton_qty"])),
                        format_number(float(shipped_metrics["carton_qty"])),
                        clean_text(carton_qty_bk),
                    ),
                    ("Size thùng", order_size_xyz, shipped_size_xyz, bang_ke_size_xyz_carton),
                ]
                if has_pallet:
                    checks.extend(
                        [
                            (
                                "Số thùng (pallet)",
                                format_number(float(order_metrics["carton_qty"])),
                                format_number(float(shipped_metrics["carton_qty"])),
                                clean_text(pallet_qty_bk),
                            ),
                            ("Size pallet", order_size_xyz, shipped_size_xyz, "120.100.h"),
                        ]
                    )
                for field_name, ol_value, shipped_value, bk_value in checks:
                    current_compare_target = field_name
                    detail_msgs: list[str] = []
                    ol_ok = True
                    shipped_ok = True
                    shipped_missing = shipped_group.empty
                    is_order_qty_metric = field_name == "Số lượng đơn hàng"
                    is_box_qty_metric = field_name.startswith("Số thùng")
                    is_qty_metric = is_order_qty_metric or is_box_qty_metric
                    if is_qty_metric:
                        if is_order_qty_metric:
                            total_value = sum_qty_total
                            # Chỉ tiêu này cho phép BK lớn hơn tổng (OL+Shipped) tối đa 1.
                            status = qty_status(format_number(total_value), bk_value)
                        else:
                            total_value = sum_carton_qty
                            # Số thùng carton/pallet so nghiêm ngặt, lệch 1 vẫn là lệch.
                            status = qty_status_strict(format_number(total_value), bk_value)
                        if status == "Lệch":
                            detail_msgs.append(
                                f"Tổng(OL+Shipped)={format_number(total_value)} khác Bảng kê={clean_text(bk_value)}"
                            )
                    elif field_name == "Màu sắc":
                        expected_color = clean_text(fabric_color_name)
                        ol_ok = True if ol_missing else color_value_matches_code(
                            ol_value, fabric_color_code, self.color_alias_by_code
                        )
                        shipped_ok = True if shipped_missing else color_value_matches_code(
                            shipped_value, fabric_color_code, self.color_alias_by_code
                        )
                        status = "Đúng" if (ol_ok and shipped_ok) else "Lệch"
                        if ol_missing:
                            detail_msgs.append("OL: Hàng đã đi, OL không có dữ liệu.")
                        elif not ol_ok:
                            detail_msgs.append(f"OL lệch màu: {clean_text(ol_value)} (kỳ vọng: {expected_color})")
                        if not shipped_missing and not shipped_ok:
                            detail_msgs.append(
                                f"Shipped lệch màu: {clean_text(shipped_value)} (kỳ vọng: {expected_color})"
                            )
                    elif field_name == "Size pallet":
                        if has_pallet:
                            status = "Lệch"
                            detail_msgs.append("Bảng kê có dòng pallet (theo rule hiện tại coi là lệch).")
                        else:
                            status = "Đúng"
                    else:
                        ol_ok = True if ol_missing else almost_equal(ol_value, bk_value)
                        shipped_ok = True if shipped_missing else almost_equal(shipped_value, bk_value)
                        status = "Đúng" if (ol_ok and shipped_ok) else "Lệch"
                        if ol_missing:
                            detail_msgs.append("OL: Hàng đã đi, OL không có dữ liệu.")
                        elif not ol_ok:
                            detail_msgs.append(f"OL lệch so với Bảng kê: {clean_text(ol_value)} vs {clean_text(bk_value)}")
                        if not shipped_missing and not shipped_ok:
                            detail_msgs.append(
                                f"Shipped lệch so với Bảng kê: {clean_text(shipped_value)} vs {clean_text(bk_value)}"
                            )

                    ol_display = "Hàng đã đi, OL không có dữ liệu" if ol_missing else clean_text(ol_value)
                    shipped_display = "Hàng chưa đi" if shipped_missing else clean_text(shipped_value)
                    if shipped_missing:
                        detail_msgs.append("Shipped: Hàng chưa đi.")
                    if is_qty_metric:
                        order_display = (
                            f"{ol_display} | Tổng: "
                            f"{format_number(sum_qty_total if field_name.startswith('Số lượng') else sum_carton_qty)}"
                        )
                    else:
                        order_display = ol_display
                    bk_display = clean_text(bk_value)
                    if detail_msgs:
                        bk_display = f"{bk_display} | " + " ; ".join(detail_msgs)
                    records.append(
                        {
                            "dg_case_no": dg_case,
                            "field_name": field_name,
                            "order_value": order_display,
                            "shipped_value": shipped_display,
                            "bang_ke_value": bk_display,
                            "auto_status": status,
                            "status_core": status,
                            "adjust_reason": " ; ".join(detail_msgs),
                            "production_no": prod_no,
                            "ship_date_display": ship_date_display,
                            "_sort_days": sort_days,
                        }
                    )
                quy_tac_extra = self._build_quy_tac_records(
                    clean_text(dg_case),
                    checks,
                    prod_no,
                    ship_date_display,
                    sort_days,
                    order_metrics,
                    ol_missing,
                    order_size_xyz,
                )
                records.extend(quy_tac_extra)
            except Exception as exc:
                append_error_record(
                    dg_case,
                    order_group if not order_group.empty else shipped_group,
                    "đối chiếu dữ liệu",
                    exc,
                    current_compare_target,
                    order_group,
                    shipped_group,
                )
                continue
        if use_progress:
            self._run_progress_done()
        result = pd.DataFrame(records)
        if result.empty:
            return pd.DataFrame(
                columns=[
                    "dg_case_no",
                    "field_name",
                    "order_value",
                    "shipped_value",
                    "bang_ke_value",
                    "auto_status",
                    "status_core",
                    "adjust_reason",
                    "production_no",
                    "ship_date_display",
                    "_sort_days",
                ]
            )
        if "_sort_days" not in result.columns:
            result["_sort_days"] = 1e15
        if "production_no" not in result.columns:
            result["production_no"] = ""
        if "ship_date_display" not in result.columns:
            result["ship_date_display"] = ""
        result = result.sort_values(["_sort_days", "dg_case_no", "field_name"], kind="stable").reset_index(drop=True)
        return result

    def save_run(
        self,
        run_type: str,
        target_dg: str | None,
        order_file: str,
        shipped_file: str,
        bang_ke_file: str,
        result_df: pd.DataFrame,
        run_label: str = "",
    ) -> int:
        t_ol = trace_orderlist_filename(order_file)
        t_ship = trace_shipped_file_mtime_local(shipped_file)
        t_a6 = trace_bang_ke_sheet_a6(bang_ke_file)
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO runs (
                run_type, target_dg, order_file, bang_ke_file, run_label,
                trace_orderlist_filename, trace_shipped_file_mtime, trace_bang_ke_a6
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_type,
                target_dg or "",
                order_file,
                bang_ke_file,
                run_label or "",
                t_ol,
                t_ship,
                t_a6,
            ),
        )
        run_id = int(cur.lastrowid)
        rows = [
            (
                run_id,
                str(rec["dg_case_no"]),
                str(rec["field_name"]),
                str(rec["order_value"]),
                str(rec.get("shipped_value", "")),
                str(rec["bang_ke_value"]),
                str(rec.get("auto_status", rec["status_core"])),
                str(rec["status_core"]),
                0,
                str(rec.get("adjust_reason", "")),
                str(rec.get("production_no", "")),
                str(rec.get("ship_date_display", "")),
            )
            for rec in result_df.to_dict("records")
        ]
        cur.executemany(
            """
            INSERT INTO run_items (
                run_id, dg_case_no, field_name, order_value, shipped_value, bang_ke_value,
                auto_status, status, is_adjusted, adjust_reason, production_no, ship_date_display
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
        conn.close()
        self.refresh_history_runs()
        return run_id

    def render_result(self, result_df: pd.DataFrame | None = None) -> None:
        """Giữ tên cũ — luôn hiển thị bảng tổng hợp theo DG."""
        self.render_check_main_table(result_df)

    def render_summary(self, result_df: pd.DataFrame | None = None) -> None:
        self.render_check_main_table(result_df)

    def on_tree_double_click(self, _event: tk.Event) -> None:
        if self.last_result_df is None or self.last_result_df.empty:
            return
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        dg_case = str(values[0]).strip()
        if not dg_case:
            return
        detail_df = self.last_result_df[self.last_result_df["dg_case_no"] == dg_case]
        if detail_df.empty:
            return
        self.open_detail_window(
            dg_case,
            detail_df,
            self.bang_ke_file_var.get().strip(),
            persist_run_id=self.current_run_id,
            meta_by_dg=self.run_case_meta_by_dg,
        )

    def _resolve_bk_path(self, path: str) -> str:
        p = clean_text(path)
        if not p:
            return ""
        try:
            return str(Path(p).resolve())
        except OSError:
            return p

    def _trim_bk_path_cache(self) -> None:
        lim = max(1, int(self._bang_ke_max_cached_files))
        while len(self._bang_ke_df_by_path) > lim:
            self._bang_ke_df_by_path.popitem(last=False)

    def _store_bang_ke_in_path_cache(self, bang_ke_file: str, bang_ke_df: pd.DataFrame) -> None:
        key = self._resolve_bk_path(bang_ke_file)
        if not key:
            return
        self._bang_ke_df_by_path[key] = bang_ke_df
        self._bang_ke_df_by_path.move_to_end(key)
        self._trim_bk_path_cache()

    def _get_bang_ke_df_cached(self, bang_ke_path: str) -> tuple[pd.DataFrame | None, str | None]:
        """Mỗi đường dẫn Bảng kê chỉ đọc đĩa tối đa 1 lần (LRU, tối đa vài file)."""
        path = clean_text(bang_ke_path)
        if not path:
            return None, "Không có đường dẫn Bảng kê."
        if not Path(path).exists():
            return None, "File Bảng kê không tồn tại hoặc không truy cập được."
        key = self._resolve_bk_path(path)
        if key in self._bang_ke_df_by_path:
            self._bang_ke_df_by_path.move_to_end(key)
            return self._bang_ke_df_by_path[key], None
        try:
            header_row = find_bang_ke_header_row(path)
            df = pd.read_excel(path, sheet_name=0, header=header_row)
            df = annotate_bang_ke_for_fast_lookup(df)
        except Exception as exc:
            return None, f"Không đọc được Bảng kê: {exc}"
        self._bang_ke_df_by_path[key] = df
        self._bang_ke_df_by_path.move_to_end(key)
        self._trim_bk_path_cache()
        return df, None

    def _npl_rows_for_detail(
        self, bang_ke_path: str, dg_case: str
    ) -> tuple[list[tuple[str, str, str, str, str]], str | None]:
        df, err = self._get_bang_ke_df_cached(bang_ke_path)
        if err is not None or df is None:
            return [], err or "Không đọc được Bảng kê."
        sub = filter_bang_ke_rows_for_dg(df, dg_case)
        return build_npl_rows_from_bk_subset(sub)

    def _upsert_case_meta(self, run_id: int, dg_case: str, user_conclusion: str, case_note: str) -> None:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO run_case_meta (run_id, dg_case_no, user_conclusion, case_note, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(run_id, dg_case_no) DO UPDATE SET
                user_conclusion = excluded.user_conclusion,
                case_note = excluded.case_note,
                updated_at = CURRENT_TIMESTAMP
            """,
            (run_id, dg_case, user_conclusion, case_note),
        )
        conn.commit()
        conn.close()

    def open_detail_window(
        self,
        dg_case: str,
        detail_df: pd.DataFrame,
        bang_ke_path: str | None = None,
        *,
        persist_run_id: int | None = None,
        meta_by_dg: dict[str, dict[str, str]] | None = None,
    ) -> None:
        win = tk.Toplevel(self.root)
        win.title(f"Chi tiết {dg_case}")
        win.geometry("1140x760")
        self.detail_win = win
        self.detail_dg_case = dg_case

        def _on_detail_close() -> None:
            self.detail_win = None
            self.detail_tree = None
            self.detail_dg_case = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_detail_close)

        outer = ttk.Frame(win)
        outer.pack(fill="both", expand=True)

        top_wrap = ttk.Frame(outer)
        top_wrap.pack(fill="both", expand=True, padx=0, pady=(10, 0))

        cols = ("field_name", "order_value", "shipped_value", "bang_ke_value", "status")
        tree = ttk.Treeview(top_wrap, columns=cols, show="headings")
        self.detail_tree = tree
        for col, width in [
            ("field_name", 180),
            ("order_value", 220),
            ("shipped_value", 220),
            ("bang_ke_value", 260),
            ("status", 360),
        ]:
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor="center")
        tree.pack(side="left", fill="both", expand=True, padx=(10, 0))
        scrollbar = ttk.Scrollbar(top_wrap, orient="vertical", command=tree.yview)
        scrollbar.pack(side="right", fill="y", padx=(0, 10))
        tree.configure(yscrollcommand=scrollbar.set)
        tree.tag_configure("ok", background="#d7f5dd")
        tree.tag_configure("bad", background="#f9d8d8")
        tree.tag_configure("rule_bad", background="#ffcc80")
        tree.bind("<Double-1>", lambda _e: self.open_row_detail_view(tree, cols))

        for rec in detail_df.to_dict("records"):
            fn = clean_text(rec.get("field_name", ""))
            is_rule = fn.startswith(RULE_CHECK_FIELD_PREFIX)
            core_st = str(rec["status_core"])
            if is_rule and core_st == "Lệch":
                disp = format_status_display(CASE_CONCLUSION_RULE_BAD, str(rec.get("adjust_reason", "")))
            else:
                disp = format_status_display(core_st, str(rec.get("adjust_reason", "")))
            if core_st == "Đúng":
                tag = "ok"
            elif is_rule:
                tag = "rule_bad"
            else:
                tag = "bad"
            tree.insert(
                "",
                "end",
                values=(rec["field_name"], rec["order_value"], rec.get("shipped_value", ""), rec["bang_ke_value"], disp),
                tags=(tag,),
            )

        meta_root = meta_by_dg if meta_by_dg is not None else self.run_case_meta_by_dg
        saved = meta_root.get(dg_case, {})
        machine = machine_case_conclusion_from_rows(detail_df)
        saved_u = clean_text(saved.get("user_conclusion", ""))
        if saved_u in (USER_CONFIRM_OK, USER_CONFIRM_BAD):
            initial_choice = saved_u
        else:
            initial_choice = USER_CONFIRM_OK if machine == "Đúng" else USER_CONFIRM_BAD
        note_init = clean_text(saved.get("case_note", ""))
        if not note_init:
            note_init = machine

        confirm = ttk.LabelFrame(outer, text="Xác nhận kết quả cho case (lưu vào run)")
        confirm.pack(fill="x", padx=10, pady=(8, 6))
        conc_var = tk.StringVar(value=initial_choice)
        ttk.Radiobutton(confirm, text=USER_CONFIRM_OK, variable=conc_var, value=USER_CONFIRM_OK).pack(anchor="w", padx=8, pady=2)
        ttk.Radiobutton(confirm, text=USER_CONFIRM_BAD, variable=conc_var, value=USER_CONFIRM_BAD).pack(anchor="w", padx=8, pady=2)
        ttk.Label(confirm, text="Ghi chú:").pack(anchor="w", padx=8, pady=(6, 0))
        note_widget = tk.Text(confirm, height=3, wrap="word")
        note_widget.pack(fill="x", padx=8, pady=(2, 6))
        note_widget.insert("1.0", note_init)

        btn_row = ttk.Frame(confirm)
        btn_row.pack(fill="x", padx=8, pady=(0, 6))

        def _on_ok() -> None:
            if persist_run_id is None:
                messagebox.showwarning("Không lưu được", "Không có run_id — hãy chạy Run hoặc mở từ Lịch sử.")
                return
            choice = conc_var.get().strip()
            if choice not in (USER_CONFIRM_OK, USER_CONFIRM_BAD):
                messagebox.showwarning("Chọn trạng thái", f"Chọn «{USER_CONFIRM_OK}» hoặc «{USER_CONFIRM_BAD}».")
                return
            note = note_widget.get("1.0", "end").strip()
            self._upsert_case_meta(persist_run_id, dg_case, choice, note)
            if self.current_run_id is not None and persist_run_id == self.current_run_id:
                self.run_case_meta_by_dg[dg_case] = {"user_conclusion": choice, "case_note": note}
                self.render_check_main_table()
            if self.history_current_run_id is not None and persist_run_id == self.history_current_run_id:
                self.history_case_meta_by_dg[dg_case] = {"user_conclusion": choice, "case_note": note}
                if self.history_last_df is not None and not self.history_last_df.empty:
                    self._render_history_items_from_df(self.history_last_df, self.history_current_run_type or "all")
            self.refresh_history_runs()
            _on_detail_close()

        ttk.Button(btn_row, text="OK — lưu và đóng", command=_on_ok).pack(side="left", padx=(0, 8))
        ttk.Button(btn_row, text="Hủy", command=_on_detail_close).pack(side="left")

        bk_path = clean_text(bang_ke_path or "") or self.bang_ke_file_var.get().strip()
        ttk.Label(
            outer,
            text="Bảng kê — NPL / định mức (cùng DG)",
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w", padx=10, pady=(10, 4))
        ttk.Label(
            outer,
            text="Không tự tải khi mở cửa sổ — bấm nút bên dưới khi cần. Mỗi file Bảng kê chỉ đọc 1 lần, các lần sau dùng cache (kể cả từ Lịch sử).",
            foreground="#555555",
            wraplength=1050,
        ).pack(anchor="w", padx=10, pady=(0, 4))
        npl_cols = ("ma_npl", "ten_npl", "mota", "dvt", "soluong")
        load_btn = ttk.Button(outer, text="Tải phụ liệu từ Bảng kê")
        load_btn.pack(anchor="w", padx=10, pady=(0, 6))
        npl_wrap = ttk.Frame(outer)
        npl_wrap.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        npl_tree = ttk.Treeview(npl_wrap, columns=npl_cols, show="headings", height=11)
        npl_head = {
            "ma_npl": "Mã NPL",
            "ten_npl": "Tên NPL",
            "mota": "Mô tả",
            "dvt": "ĐVT (cột N)",
            "soluong": "Số lượng (cột P)",
        }
        for col, width in [
            ("ma_npl", 120),
            ("ten_npl", 200),
            ("mota", 240),
            ("dvt", 100),
            ("soluong", 100),
        ]:
            npl_tree.heading(col, text=npl_head[col])
            npl_tree.column(col, width=width, anchor="center")
        npl_sb = ttk.Scrollbar(npl_wrap, orient="vertical", command=npl_tree.yview)
        npl_tree.configure(yscrollcommand=npl_sb.set)
        npl_tree.pack(side="left", fill="both", expand=True)
        npl_sb.pack(side="right", fill="y")

        def _on_load_npl() -> None:
            try:
                if not win.winfo_exists():
                    return
            except tk.TclError:
                return
            load_btn.configure(state="disabled", text="Đang đọc…")
            win.update_idletasks()
            self.root.update_idletasks()
            for item in npl_tree.get_children():
                npl_tree.delete(item)
            npl_rows, npl_err = self._npl_rows_for_detail(bk_path, dg_case)
            try:
                if win.winfo_exists():
                    load_btn.configure(state="normal", text="Tải lại phụ liệu")
            except tk.TclError:
                return
            if npl_err:
                npl_tree.insert("", "end", values=(npl_err, "", "", "", ""))
            else:
                for tup in npl_rows:
                    npl_tree.insert("", "end", values=tup)

        if not bk_path:
            load_btn.configure(state="disabled", text="Thiếu file Bảng kê")
        else:
            load_btn.configure(command=_on_load_npl)

    def open_row_detail_view(self, tree: ttk.Treeview, columns: tuple[str, ...]) -> None:
        selected = tree.selection()
        if not selected:
            return
        values = tree.item(selected[0], "values")
        if not values:
            return
        win = tk.Toplevel(self.root)
        win.title("Chi tiết dòng")
        win.geometry("860x420")
        text = tk.Text(win, wrap="word")
        text.pack(fill="both", expand=True, padx=10, pady=10)
        for idx, col in enumerate(columns):
            val = str(values[idx]) if idx < len(values) else ""
            text.insert("end", f"{col}:\n{val}\n\n")
        text.configure(state="disabled")


def main(back_to_launcher: callable | None = None) -> None:
    root = tk.Tk()
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass
    app = OrderlistCheckerApp(root, back_to_launcher=back_to_launcher)
    root.mainloop()


if __name__ == "__main__":
    main()
