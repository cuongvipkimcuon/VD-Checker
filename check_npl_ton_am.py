import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
from pathlib import Path
import sqlite3
import unicodedata
import json
import hashlib
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill


DB_FILE = "npl_checker.db"
CONFIG_FILE = "npl_checker_config.json"
RESULT_COLUMNS = [
    "so_o",
    "ma_npl",
    "ten_npl",
    "hang_nhap",
    "ton_thuc_te",
    "ton_dm_chua_xuat",
    "so_luong_can",
    "tan_dung",
    "ket_luan",
]


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.replace("đ", "d")


def find_header_row(file_path: str, must_have_keywords: list[str]) -> int:
    preview = pd.read_excel(file_path, sheet_name=0, header=None, nrows=40)
    for idx, row in preview.iterrows():
        row_text = " | ".join(
            normalize_text(cell) for cell in row.tolist() if normalize_text(cell)
        )
        if all(keyword in row_text for keyword in must_have_keywords):
            return int(idx)
    raise ValueError(f"Khong tim thay dong header phu hop trong file: {file_path}")


def find_column(df: pd.DataFrame, expected_keywords: list[str]) -> str:
    col_map = {col: normalize_text(col) for col in df.columns}
    for col, col_text in col_map.items():
        if all(keyword in col_text for keyword in expected_keywords):
            return col
    raise ValueError(f"Khong tim thay cot voi tu khoa: {expected_keywords}")


def init_db() -> None:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_name TEXT,
            so_o TEXT NOT NULL,
            bom_file TEXT,
            stock_file TEXT,
            note TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS run_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL,
            so_o TEXT,
            ma_npl TEXT,
            ten_npl TEXT,
            hang_nhap TEXT,
            ton_thuc_te REAL,
            ton_dm_chua_xuat REAL,
            ket_luan TEXT,
            FOREIGN KEY (run_id) REFERENCES runs(id) ON DELETE CASCADE
        )
        """
    )
    existing_runs_cols = {row[1] for row in cur.execute("PRAGMA table_info(runs)").fetchall()}
    if "trace_bom_a6" not in existing_runs_cols:
        cur.execute("ALTER TABLE runs ADD COLUMN trace_bom_a6 TEXT DEFAULT ''")
    if "trace_stock_a6" not in existing_runs_cols:
        cur.execute("ALTER TABLE runs ADD COLUMN trace_stock_a6 TEXT DEFAULT ''")
    existing_item_cols = {row[1] for row in cur.execute("PRAGMA table_info(run_items)").fetchall()}
    if "hang_nhap" not in existing_item_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN hang_nhap TEXT DEFAULT ''")
    if "so_luong_can" not in existing_item_cols:
        cur.execute("ALTER TABLE run_items ADD COLUMN so_luong_can REAL")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS npl_utilization_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            so_o TEXT NOT NULL,
            ma_npl TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(so_o, ma_npl)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS file_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_kind TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            file_path TEXT DEFAULT '',
            payload_json TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(file_kind, file_hash)
        )
        """
    )
    conn.commit()
    conn.close()


def file_sha256(file_path: str) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def classify_stock_status(ton_thuc_te: float | None, ton_dm: float | None, so_luong_can: float) -> str:
    if pd.isna(ton_thuc_te) and pd.isna(ton_dm):
        return "Khong tim thay ma trong ton kho"
    if (pd.notna(ton_thuc_te) and float(ton_thuc_te) < 0) or (pd.notna(ton_dm) and float(ton_dm) < 0):
        return "Thieu NPL"
    if pd.notna(ton_thuc_te) and pd.notna(ton_dm) and float(ton_thuc_te) > 0 and float(ton_dm) > 0:
        ton_tt_sau_tru = float(ton_thuc_te) - float(so_luong_can)
        ton_dm_sau_tru = float(ton_dm) - float(so_luong_can)
        if ton_tt_sau_tru >= 0 and ton_dm_sau_tru >= 0:
            return "An toan"
        return "Chu y"
    return "Chu y"


def classify_stock_status_series(
    ton_thuc_te: pd.Series,
    ton_dm: pd.Series,
    so_luong_can: pd.Series,
) -> pd.Series:
    ton_tt = pd.to_numeric(ton_thuc_te, errors="coerce")
    ton_dm_num = pd.to_numeric(ton_dm, errors="coerce")
    slc = pd.to_numeric(so_luong_can, errors="coerce").fillna(0.0)

    no_stock = ton_tt.isna() & ton_dm_num.isna()
    thieu = ((ton_tt.notna()) & (ton_tt < 0)) | ((ton_dm_num.notna()) & (ton_dm_num < 0))
    both_pos = ton_tt.notna() & ton_dm_num.notna() & (ton_tt > 0) & (ton_dm_num > 0)
    an_toan = both_pos & ((ton_tt - slc) >= 0) & ((ton_dm_num - slc) >= 0)
    chu_y = ~(no_stock | thieu | an_toan)
    return pd.Series(
        np.select(
            [no_stock, thieu, an_toan, chu_y],
            ["Khong tim thay ma trong ton kho", "Thieu NPL", "An toan", "Chu y"],
            default="Chu y",
        ),
        index=ton_tt.index,
    )


def ket_luan_tag(ket_luan: str) -> str:
    if ket_luan == "Thieu NPL":
        return "negative"
    if ket_luan == "Chu y":
        return "warning"
    return "normal"


def normalize_ma_series(series: pd.Series) -> pd.Series:
    return series.map(normalize_text)


def build_hang_nhap_label_by_ma(
    ma_series: pd.Series,
    hang_nhap_series: pd.Series,
    date_series: pd.Series | None = None,
    max_recent_rows_per_ma: int = 20,
    focus_ma_norm_series: pd.Series | None = None,
) -> dict[str, str]:
    tmp_dict: dict[str, pd.Series] = {
        "ma_norm": normalize_ma_series(ma_series),
        "hang_nhap_text": hang_nhap_series.fillna("").astype(str).str.strip(),
    }
    if date_series is not None:
        tmp_dict["date_raw"] = date_series
    tmp = pd.DataFrame(tmp_dict)
    if "date_raw" in tmp.columns:
        tmp["date_value"] = pd.to_datetime(tmp["date_raw"], errors="coerce", dayfirst=True)
    tmp = tmp[tmp["ma_norm"] != ""].copy()
    if tmp.empty:
        return {}
    if focus_ma_norm_series is not None:
        focus_keys = {normalize_text(v) for v in focus_ma_norm_series.tolist()}
        focus_keys.discard("")
    else:
        focus_keys = set(tmp["ma_norm"].tolist())
    if not focus_keys:
        return {}
    tmp = tmp[tmp["ma_norm"].isin(focus_keys)].copy()
    if tmp.empty:
        return {}
    if "date_value" in tmp.columns and tmp["date_value"].notna().any():
        sampled = (
            tmp.sort_values(["ma_norm", "date_value"], ascending=[True, False])
            .groupby("ma_norm", as_index=False, sort=False)
            .head(max_recent_rows_per_ma)
            .copy()
        )
    else:
        sampled = (
            tmp.groupby("ma_norm", as_index=False, sort=False)
            .tail(max_recent_rows_per_ma)
            .copy()
        )
    sampled["is_d"] = sampled["hang_nhap_text"].str.upper().str.contains("D", na=False).astype(float)
    sampled["rank_recent"] = sampled.groupby("ma_norm", sort=False).cumcount()
    sampled["base_weight"] = np.select(
        [
            sampled["rank_recent"] <= 4,
            (sampled["rank_recent"] >= 5) & (sampled["rank_recent"] <= 14),
            sampled["rank_recent"] >= 15,
        ],
        [0.12, 0.03, 0.02],
        default=0.0,
    )
    sampled["weight_sum_per_ma"] = sampled.groupby("ma_norm", sort=False)["base_weight"].transform("sum")
    sampled["norm_weight"] = np.where(
        sampled["weight_sum_per_ma"] > 0,
        sampled["base_weight"] / sampled["weight_sum_per_ma"],
        0.0,
    )
    sampled["weighted_d"] = sampled["is_d"] * sampled["norm_weight"]
    stats = sampled.groupby("ma_norm", as_index=False).agg(
        weighted_d=("weighted_d", "sum"),
        row_count=("ma_norm", "size"),
    )
    labels: dict[str, str] = {}
    for _, r in stats.iterrows():
        total = int(r["row_count"])
        if total <= 0:
            labels[str(r["ma_norm"])] = ""
            continue
        pct_d = float(r["weighted_d"]) * 100.0
        if pct_d >= 50:
            labels[str(r["ma_norm"])] = f"Hang nhap ({pct_d:.0f}%)"
        else:
            labels[str(r["ma_norm"])] = f"Hang noi ({100.0 - pct_d:.0f}%)"
    return labels


def select_hang_nhap_label(labels: pd.Series) -> str:
    texts = [str(x).strip() for x in labels.tolist() if str(x).strip()]
    if not texts:
        return ""
    best = texts[0]
    best_pct = -1.0
    for text in texts:
        l = text.rfind("(")
        r = text.rfind("%")
        pct = -1.0
        if l >= 0 and r > l:
            try:
                pct = float(text[l + 1 : r].strip())
            except ValueError:
                pct = -1.0
        if pct > best_pct:
            best_pct = pct
            best = text
    return best


def summarize_tan_dung(series: pd.Series) -> str:
    vals = {normalize_text(v) for v in series.tolist()}
    vals.discard("")
    has_yes = "co" in vals
    has_no = "khong" in vals
    if has_yes and has_no:
        return "Mot phan"
    if has_yes:
        return "Co"
    return "Khong"


def load_config() -> dict:
    config_path = Path(CONFIG_FILE)
    if not config_path.exists():
        return {}
    try:
        return json.loads(config_path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_config(data: dict) -> None:
    config_path = Path(CONFIG_FILE)
    config_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def trace_sheet_a6(file_path: str) -> str:
    p = str(file_path or "").strip()
    if not p:
        return ""
    try:
        df = pd.read_excel(p, sheet_name=0, header=None, usecols=[0], nrows=6)
        if df.shape[0] <= 5:
            return ""
        v = df.iloc[5, 0]
        if pd.isna(v):
            return ""
        return str(v).strip()
    except Exception:
        return ""


class App:
    def __init__(self, root: tk.Tk, back_to_launcher: callable | None = None):
        init_db()
        self.root = root
        self.back_to_launcher = back_to_launcher
        self.root.title("NPL Checker - Kiem tra ton am + database")
        self.root.geometry("1320x760")

        self.bom_file_var = tk.StringVar()
        self.stock_file_var = tk.StringVar()
        self.o_number_var = tk.StringVar()
        self.search_result_var = tk.StringVar()
        self.search_db_var = tk.StringVar()

        self.last_result_df: pd.DataFrame | None = None
        self.last_result_base_df: pd.DataFrame | None = None
        self.last_summary_df: pd.DataFrame | None = None
        self.last_summary_source_df: pd.DataFrame | None = None
        self.last_summary_run_ids: list[int] = []
        self.last_trace_bom_a6 = ""
        self.last_trace_stock_a6 = ""
        self.loaded_run_id: int | None = None
        self.last_cache_info = ""
        self.config = load_config()

        self._build_ui()
        self._load_last_paths()
        self.refresh_runs()

    def _build_ui(self) -> None:
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True)

        self.tab_check = ttk.Frame(self.notebook, padding=10)
        self.tab_db = ttk.Frame(self.notebook, padding=10)
        self.tab_summary = ttk.Frame(self.notebook, padding=10)
        self.tab_rules = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.tab_check, text="Fast Check")
        self.notebook.add(self.tab_db, text="Data")
        self.notebook.add(self.tab_summary, text="Super Report")
        self.notebook.add(self.tab_rules, text="Quy tac Tan dung")

        self._build_tab_check()
        self._build_tab_db()
        self._build_tab_summary()
        self._build_tab_rules()

    def _build_tab_check(self) -> None:
        top = ttk.Frame(self.tab_check)
        top.pack(fill="x")

        ttk.Label(top, text="File bang ke dinh muc:").grid(
            row=0, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(top, textvariable=self.bom_file_var, width=95).grid(
            row=0, column=1, padx=(0, 8), pady=4, sticky="ew"
        )
        ttk.Button(top, text="Chon file", command=self.choose_bom_file).grid(row=0, column=2)

        ttk.Label(top, text="File tong hop nhap xuat ton:").grid(
            row=1, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(top, textvariable=self.stock_file_var, width=95).grid(
            row=1, column=1, padx=(0, 8), pady=4, sticky="ew"
        )
        ttk.Button(top, text="Chon file", command=self.choose_stock_file).grid(row=1, column=2)

        ttk.Label(top, text="So O can check:").grid(
            row=2, column=0, sticky="w", padx=(0, 8), pady=8
        )
        ttk.Entry(top, textvariable=self.o_number_var, width=30).grid(row=2, column=1, sticky="w")
        action_row = ttk.Frame(top)
        action_row.grid(row=2, column=2, sticky="e")
        ttk.Button(action_row, text="Kiem tra", command=self.run_check).pack(side="left")
        ttk.Button(action_row, text="Xuat Excel", command=self.export_excel).pack(side="left", padx=(6, 0))

        ttk.Label(top, text="Search ket qua:").grid(row=3, column=0, sticky="w", pady=(8, 0))
        ttk.Entry(top, textvariable=self.search_result_var, width=40).grid(
            row=3, column=1, sticky="w", pady=(8, 0)
        )
        filter_row = ttk.Frame(top)
        filter_row.grid(row=3, column=2, sticky="e", pady=(8, 0))
        ttk.Button(filter_row, text="Loc", command=self.filter_current_results).pack(side="left")
        ttk.Button(filter_row, text="Bo loc", command=self.clear_result_filter).pack(side="left", padx=(6, 0))
        top.columnconfigure(1, weight=1)

        self.status_label = ttk.Label(
            self.tab_check, text="San sang.", foreground="#1f4e79", padding=(0, 8)
        )
        self.status_label.pack(fill="x")

        self.result_tree = self._create_result_tree(self.tab_check)

        action_wrap = ttk.Frame(self.tab_check)
        action_wrap.pack(fill="x", pady=(8, 0))
        ttk.Button(action_wrap, text="Save vao database", command=self.save_current_run).pack(side="left")
        ttk.Button(
            action_wrap,
            text="Tan dung NPL da chon",
            command=self.add_utilization_from_selected,
        ).pack(side="left", padx=6)
        ttk.Button(
            action_wrap,
            text="Bo Tan dung NPL da chon",
            command=self.deactivate_utilization_from_selected,
        ).pack(side="left")
        ttk.Button(action_wrap, text="Mo tab tong hop", command=self.switch_to_summary).pack(side="left", padx=6)
        if self.back_to_launcher is not None:
            ttk.Button(action_wrap, text="Back ve Launcher", command=self.go_back_to_launcher).pack(
                side="left", padx=6
            )

    def go_back_to_launcher(self) -> None:
        if self.back_to_launcher is None:
            return
        self.root.destroy()
        self.back_to_launcher()

    def _build_tab_db(self) -> None:
        top = ttk.Frame(self.tab_db)
        top.pack(fill="x")
        ttk.Label(top, text="Search run (so O / ten run / ghi chu):").pack(side="left")
        ttk.Entry(top, textvariable=self.search_db_var, width=45).pack(side="left", padx=6)
        ttk.Button(top, text="Search", command=self.refresh_runs).pack(side="left")
        ttk.Button(top, text="Refresh", command=self.refresh_runs).pack(side="left", padx=6)

        run_columns = ("id", "run_name", "so_o", "created_at", "note", "item_count")
        self.runs_tree = ttk.Treeview(self.tab_db, columns=run_columns, show="headings", height=12)
        for col, width in [
            ("id", 70),
            ("run_name", 220),
            ("so_o", 120),
            ("created_at", 160),
            ("note", 360),
            ("item_count", 90),
        ]:
            self.runs_tree.heading(col, text=col)
            self.runs_tree.column(col, width=width, anchor="w")
        self.runs_tree.pack(fill="x", pady=(8, 8))

        btn_wrap = ttk.Frame(self.tab_db)
        btn_wrap.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_wrap, text="Load run", command=self.load_selected_run).pack(side="left")
        ttk.Button(btn_wrap, text="Update note", command=self.update_selected_run_note).pack(side="left", padx=6)
        ttk.Button(btn_wrap, text="Delete run", command=self.delete_selected_run).pack(side="left")
        ttk.Button(btn_wrap, text="Export run", command=self.export_selected_run).pack(side="left", padx=6)
        ttk.Button(btn_wrap, text="Import run", command=self.import_run_from_file).pack(side="left")

        ttk.Label(self.tab_db, text="Du lieu chi tiet run da chon:").pack(anchor="w")
        self.db_item_tree = self._create_result_tree(self.tab_db)

    def _build_tab_summary(self) -> None:
        top = ttk.Frame(self.tab_summary)
        top.pack(fill="x")
        ttk.Button(
            top,
            text="Tao bao cao tu run da chon",
            command=self.build_summary_report,
        ).pack(side="left")
        ttk.Button(top, text="Refresh", command=self.build_summary_report).pack(side="left", padx=6)
        ttk.Button(top, text="Export Excel", command=self.export_summary_excel).pack(side="left", padx=6)

        run_picker = ttk.Frame(self.tab_summary)
        run_picker.pack(fill="x", pady=(8, 0))
        ttk.Label(run_picker, text="Chon run de gom (giu Ctrl/Shift de chon nhieu):").pack(
            anchor="w"
        )
        list_wrap = ttk.Frame(run_picker)
        list_wrap.pack(fill="x", pady=(4, 0))
        self.summary_runs_listbox = tk.Listbox(
            list_wrap,
            selectmode="extended",
            height=6,
            exportselection=False,
        )
        self.summary_runs_listbox.pack(side="left", fill="x", expand=True)
        summary_list_scroll = ttk.Scrollbar(
            list_wrap, orient="vertical", command=self.summary_runs_listbox.yview
        )
        self.summary_runs_listbox.configure(yscrollcommand=summary_list_scroll.set)
        summary_list_scroll.pack(side="right", fill="y")

        picker_actions = ttk.Frame(run_picker)
        picker_actions.pack(fill="x", pady=(4, 0))
        ttk.Button(
            picker_actions, text="Load danh sach run", command=self.refresh_summary_run_list
        ).pack(side="left")
        ttk.Button(
            picker_actions, text="Chon tat ca", command=self.select_all_summary_runs
        ).pack(side="left", padx=6)
        ttk.Button(
            picker_actions, text="Bo chon", command=self.clear_summary_run_selection
        ).pack(side="left")

        self.summary_label = ttk.Label(
            self.tab_summary, text="Bao cao tong hop chua duoc tao.", padding=(0, 8)
        )
        self.summary_label.pack(fill="x")

        cols = (
            "ma_npl",
            "ten_npl",
            "hang_nhap",
            "so_o",
            "run_id",
            "ton_thuc_te",
            "ton_dm_chua_xuat",
            "so_luong_can",
            "tan_dung",
            "ket_luan",
        )
        self.summary_tree = ttk.Treeview(self.tab_summary, columns=cols, show="headings")
        for name, width, anchor in [
            ("ma_npl", 150, "center"),
            ("ten_npl", 420, "w"),
            ("hang_nhap", 120, "center"),
            ("so_o", 120, "center"),
            ("run_id", 80, "center"),
            ("ton_thuc_te", 140, "center"),
            ("ton_dm_chua_xuat", 170, "center"),
            ("so_luong_can", 130, "center"),
            ("tan_dung", 100, "center"),
            ("ket_luan", 180, "center"),
        ]:
            self.summary_tree.heading(name, text=name)
            self.summary_tree.column(name, width=width, anchor=anchor)
        self.summary_tree.tag_configure("negative", background="#ffd9d9")
        self.summary_tree.tag_configure("warning", background="#ffe6bf")
        self.summary_tree.tag_configure("normal", background="#e7f7e7")

        wrap = ttk.Frame(self.tab_summary)
        wrap.pack(fill="both", expand=True)
        y_scroll = ttk.Scrollbar(wrap, orient="vertical", command=self.summary_tree.yview)
        self.summary_tree.configure(yscrollcommand=y_scroll.set)
        self.summary_tree.pack(side="left", fill="both", expand=True)
        y_scroll.pack(side="right", fill="y")

    def _build_tab_rules(self) -> None:
        top = ttk.Frame(self.tab_rules)
        top.pack(fill="x")
        ttk.Label(
            top,
            text="Quy tac Tan dung theo So O + Ma NPL. Active = khi tinh ton am, so luong can cua NPL se tinh = 0.",
            wraplength=1180,
        ).pack(anchor="w")

        btns = ttk.Frame(self.tab_rules)
        btns.pack(fill="x", pady=(8, 8))
        ttk.Button(btns, text="Them moi", command=self.add_utilization_manual).pack(side="left")
        ttk.Button(btns, text="UnActive", command=self.deactivate_selected_rule).pack(side="left", padx=6)
        ttk.Button(btns, text="Kich hoat lai", command=self.activate_selected_rule).pack(side="left")
        ttk.Button(btns, text="Refresh", command=self.refresh_utilization_rules).pack(side="left", padx=12)

        cols = ("so_o", "ma_npl", "active", "created_at", "updated_at")
        self.rules_tree = ttk.Treeview(self.tab_rules, columns=cols, show="headings", height=20)
        for name, text, width in [
            ("so_o", "So O", 180),
            ("ma_npl", "Ma NPL", 180),
            ("active", "Kich hoat", 90),
            ("created_at", "Ngay tao", 180),
            ("updated_at", "Cap nhat", 180),
        ]:
            self.rules_tree.heading(name, text=text)
            self.rules_tree.column(name, width=width, anchor="center")
        self.rules_tree.pack(fill="both", expand=True)
        self.refresh_utilization_rules()

    def _create_result_tree(self, parent: ttk.Frame) -> ttk.Treeview:
        wrap = ttk.Frame(parent)
        wrap.pack(fill="both", expand=True)

        columns = tuple(RESULT_COLUMNS)
        tree = ttk.Treeview(wrap, columns=columns, show="headings")
        tree.heading("so_o", text="So O")
        tree.heading("ma_npl", text="Ma NPL")
        tree.heading("ten_npl", text="Ten NPL")
        tree.heading("hang_nhap", text="Hang nhap")
        tree.heading("ton_thuc_te", text="Ton thuc te")
        tree.heading("ton_dm_chua_xuat", text="Ton - dinh muc chua xuat")
        tree.heading("so_luong_can", text="So luong can")
        tree.heading("tan_dung", text="Tan dung")
        tree.heading("ket_luan", text="Ket luan")
        tree.column("so_o", width=140, anchor="center")
        tree.column("ma_npl", width=150, anchor="center")
        tree.column("ten_npl", width=400, anchor="w")
        tree.column("hang_nhap", width=110, anchor="center")
        tree.column("ton_thuc_te", width=120, anchor="center")
        tree.column("ton_dm_chua_xuat", width=160, anchor="center")
        tree.column("so_luong_can", width=120, anchor="center")
        tree.column("tan_dung", width=90, anchor="center")
        tree.column("ket_luan", width=220, anchor="center")
        tree.tag_configure("negative", background="#ffd9d9")
        tree.tag_configure("warning", background="#ffe6bf")
        tree.tag_configure("normal", background="#e7f7e7")

        y_scroll = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)
        return tree

    def choose_bom_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Chon file Bang ke dinh muc",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.bom_file_var.set(path)
            self.last_trace_bom_a6 = ""
            self._save_last_paths()

    def choose_stock_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Chon file Tong hop nhap xuat ton",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.stock_file_var.set(path)
            self.last_trace_stock_a6 = ""
            self._save_last_paths()

    def _load_last_paths(self) -> None:
        bom = self.config.get("last_bom_file", "")
        stock = self.config.get("last_stock_file", "")
        if bom and Path(bom).exists():
            self.bom_file_var.set(bom)
        if stock and Path(stock).exists():
            self.stock_file_var.set(stock)

    def _save_last_paths(self) -> None:
        self.config["last_bom_file"] = self.bom_file_var.get().strip()
        self.config["last_stock_file"] = self.stock_file_var.get().strip()
        save_config(self.config)

    def _load_df_from_cache(self, file_kind: str, file_hash: str) -> pd.DataFrame | None:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        row = cur.execute(
            """
            SELECT payload_json
            FROM file_cache
            WHERE file_kind = ? AND file_hash = ?
            """,
            (file_kind, file_hash),
        ).fetchone()
        conn.close()
        if not row or not row[0]:
            return None
        try:
            return pd.read_json(row[0], orient="split")
        except Exception:
            return None

    def _save_df_to_cache(self, file_kind: str, file_path: str, file_hash: str, df: pd.DataFrame) -> None:
        payload = df.to_json(orient="split", force_ascii=False)
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO file_cache (file_kind, file_hash, file_path, payload_json)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(file_kind, file_hash) DO UPDATE SET
                file_path = excluded.file_path,
                payload_json = excluded.payload_json,
                updated_at = CURRENT_TIMESTAMP
            """,
            (file_kind, file_hash, file_path, payload),
        )
        conn.commit()
        conn.close()

    def _load_or_build_bom_cache(self, bom_file: str) -> tuple[pd.DataFrame, bool]:
        bom_hash = file_sha256(bom_file)
        cached = self._load_df_from_cache("bom", bom_hash)
        if cached is not None and not cached.empty:
            return cached, True
        bom_header = find_header_row(bom_file, must_have_keywords=["s/o", "ma npl", "ten npl"])
        bom_df = pd.read_excel(bom_file, sheet_name=0, header=bom_header)
        so_o_col = find_column(bom_df, ["s/o"])
        ma_npl_col = find_column(bom_df, ["ma", "npl"])
        ten_npl_col = find_column(bom_df, ["ten", "npl"])
        date_series = bom_df.iloc[:, 1] if bom_df.shape[1] > 1 else pd.Series([""] * len(bom_df), index=bom_df.index)
        hang_nhap_series = (
            bom_df.iloc[:, 12] if bom_df.shape[1] > 12 else pd.Series([""] * len(bom_df), index=bom_df.index)
        )
        so_luong_can_series = (
            bom_df.iloc[:, 15] if bom_df.shape[1] > 15 else pd.Series([0] * len(bom_df), index=bom_df.index)
        )
        out = pd.DataFrame(
            {
                "so_o": bom_df[so_o_col].astype(str),
                "so_o_norm": normalize_ma_series(bom_df[so_o_col]),
                "ma_npl": bom_df[ma_npl_col].astype(str),
                "ma_npl_norm": normalize_ma_series(bom_df[ma_npl_col]),
                "ten_npl": bom_df[ten_npl_col].astype(str),
                "date_raw": date_series,
                "hang_nhap_text": hang_nhap_series.fillna("").astype(str).str.strip(),
                "so_luong_can": pd.to_numeric(so_luong_can_series, errors="coerce").fillna(0.0),
            }
        )
        self._save_df_to_cache("bom", bom_file, bom_hash, out)
        return out, False

    def _load_or_build_stock_cache(self, stock_file: str) -> tuple[pd.DataFrame, bool]:
        stock_hash = file_sha256(stock_file)
        cached = self._load_df_from_cache("stock", stock_hash)
        if cached is not None and not cached.empty:
            return cached, True
        stock_header = find_header_row(stock_file, must_have_keywords=["ma vat tu", "ton thuc te"])
        stock_df = pd.read_excel(stock_file, sheet_name=0, header=stock_header)
        stock_ma_col = find_column(stock_df, ["ma", "vat tu"])
        ton_thuc_te_col = find_column(stock_df, ["ton", "thuc", "te"])
        ton_dm_chua_xuat_col = find_column(stock_df, ["ton", "dinh", "muc", "chua"])
        out = pd.DataFrame(
            {
                "ma_npl_norm": normalize_ma_series(stock_df[stock_ma_col]),
                "ton_thuc_te": pd.to_numeric(stock_df[ton_thuc_te_col], errors="coerce"),
                "ton_dm_chua_xuat": pd.to_numeric(stock_df[ton_dm_chua_xuat_col], errors="coerce"),
            }
        ).drop_duplicates(subset=["ma_npl_norm"], keep="first")
        self._save_df_to_cache("stock", stock_file, stock_hash, out)
        return out, False

    def run_check(self) -> None:
        try:
            bom_file = self.bom_file_var.get().strip()
            stock_file = self.stock_file_var.get().strip()
            o_number_input = self.o_number_var.get().strip()
            if not bom_file or not stock_file or not o_number_input:
                messagebox.showwarning("Thieu du lieu", "Hay chon 2 file va nhap so O.")
                return
            if not Path(bom_file).exists() or not Path(stock_file).exists():
                messagebox.showerror("Loi file", "Duong dan file khong ton tai.")
                return

            self.status_label.config(text="Dang doc file va xu ly...")
            self.root.update_idletasks()
            self._save_last_paths()

            result_df = self.check_negative_stock(bom_file, stock_file, o_number_input)
            self.last_result_base_df = result_df.copy()
            self.last_result_df = self._apply_utilization_rules(result_df)
            self.last_trace_bom_a6 = trace_sheet_a6(bom_file)
            self.last_trace_stock_a6 = trace_sheet_a6(stock_file)
            self.loaded_run_id = None
            self.render_result(self.result_tree, self.last_result_df)
            negatives = (self.last_result_df["ket_luan"] == "Thieu NPL").sum()
            self.status_label.config(
                text=f"Xong: {len(self.last_result_df)} dong lien quan, {negatives} dong thieu NPL. {self.last_cache_info}".strip()
            )
        except Exception as exc:
            messagebox.showerror("Loi", str(exc))
            self.status_label.config(text="Co loi trong qua trinh xu ly.")

    def check_negative_stock(
        self, bom_file: str, stock_file: str, o_number_input: str
    ) -> pd.DataFrame:
        bom_df, bom_cached = self._load_or_build_bom_cache(bom_file)
        stock_df, stock_cached = self._load_or_build_stock_cache(stock_file)
        self.last_cache_info = f"(BOM: {'cache' if bom_cached else 'read file'} | STOCK: {'cache' if stock_cached else 'read file'})"
        o_lookup = normalize_text(o_number_input)
        filtered = bom_df[bom_df["so_o_norm"] == o_lookup].copy()
        if filtered.empty:
            raise ValueError(f"Khong tim thay so O: {o_number_input}")
        filtered["so_luong_can"] = pd.to_numeric(filtered["so_luong_can"], errors="coerce").fillna(0.0)
        hang_nhap_by_ma = build_hang_nhap_label_by_ma(
            bom_df["ma_npl"],
            bom_df["hang_nhap_text"],
            date_series=bom_df["date_raw"],
            max_recent_rows_per_ma=20,
            focus_ma_norm_series=filtered["ma_npl_norm"],
        )
        filtered["hang_nhap"] = filtered["ma_npl_norm"].map(hang_nhap_by_ma).fillna("")
        merged = filtered.merge(
            stock_df[["ma_npl_norm", "ton_thuc_te", "ton_dm_chua_xuat"]],
            how="left",
            left_on="ma_npl_norm",
            right_on="ma_npl_norm",
        )
        merged["ton_thuc_te"] = pd.to_numeric(merged["ton_thuc_te"], errors="coerce")
        merged["ton_dm_chua_xuat"] = pd.to_numeric(merged["ton_dm_chua_xuat"], errors="coerce")

        merged["ket_luan"] = classify_stock_status_series(
            merged["ton_thuc_te"],
            merged["ton_dm_chua_xuat"],
            merged["so_luong_can"],
        )
        return pd.DataFrame(
            {
                "so_o": merged["so_o"],
                "ma_npl": merged["ma_npl"],
                "ten_npl": merged["ten_npl"],
                "hang_nhap": merged["hang_nhap"],
                "ton_thuc_te": merged["ton_thuc_te"],
                "ton_dm_chua_xuat": merged["ton_dm_chua_xuat"],
                "so_luong_can": merged["so_luong_can"],
                "ket_luan": merged["ket_luan"],
            }
        )

    def render_result(self, tree: ttk.Treeview, result_df: pd.DataFrame) -> None:
        for row_id in tree.get_children():
            tree.delete(row_id)
        for _, row in result_df.iterrows():
            tag = ket_luan_tag(str(row["ket_luan"]))
            tree.insert(
                "",
                "end",
                values=(
                    row["so_o"],
                    row["ma_npl"],
                    row["ten_npl"],
                    row["hang_nhap"],
                    "" if pd.isna(row["ton_thuc_te"]) else f"{row['ton_thuc_te']:.4f}",
                    "" if pd.isna(row["ton_dm_chua_xuat"]) else f"{row['ton_dm_chua_xuat']:.4f}",
                    "" if pd.isna(row["so_luong_can"]) else f"{row['so_luong_can']:.4f}",
                    row.get("tan_dung", "Khong"),
                    row["ket_luan"],
                ),
                tags=(tag,),
            )

    def filter_current_results(self) -> None:
        if self.last_result_df is None:
            return
        kw = normalize_text(self.search_result_var.get())
        if not kw:
            self.render_result(self.result_tree, self.last_result_df)
            return
        mask = self.last_result_df.apply(
            lambda row: kw in normalize_text(row["so_o"])
            or kw in normalize_text(row["ma_npl"])
            or kw in normalize_text(row["ten_npl"])
            or kw in normalize_text(row.get("tan_dung", ""))
            or kw in normalize_text(row["ket_luan"]),
            axis=1,
        )
        self.render_result(self.result_tree, self.last_result_df[mask].copy())

    def clear_result_filter(self) -> None:
        self.search_result_var.set("")
        if self.last_result_df is not None:
            self.render_result(self.result_tree, self.last_result_df)

    def _active_utilization_keys(self) -> set[tuple[str, str]]:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        rows = cur.execute(
            """
            SELECT so_o, ma_npl
            FROM npl_utilization_rules
            WHERE is_active = 1
            """
        ).fetchall()
        conn.close()
        return {(normalize_text(so_o), normalize_text(ma_npl)) for so_o, ma_npl in rows}

    def _apply_utilization_rules(self, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df.copy()
        out = df.copy()
        out["tan_dung"] = "Khong"
        keys = self._active_utilization_keys()
        if keys:
            pair = list(zip(normalize_ma_series(out["so_o"]), normalize_ma_series(out["ma_npl"])))
            util_mask = pd.Series([(so_o, ma) in keys for so_o, ma in pair], index=out.index)
            out.loc[util_mask, "so_luong_can"] = 0.0
            out.loc[util_mask, "tan_dung"] = "Co"
        out["ket_luan"] = classify_stock_status_series(
            out["ton_thuc_te"], out["ton_dm_chua_xuat"], out["so_luong_can"]
        )
        return out

    def _upsert_utilization_rule(self, so_o: str, ma_npl: str, is_active: int) -> None:
        so_o_clean = str(so_o).strip()
        ma_npl_clean = str(ma_npl).strip()
        if not so_o_clean or not ma_npl_clean:
            raise ValueError("So O va Ma NPL khong duoc de trong.")
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO npl_utilization_rules (so_o, ma_npl, is_active)
            VALUES (?, ?, ?)
            ON CONFLICT(so_o, ma_npl) DO UPDATE SET
                is_active = excluded.is_active,
                updated_at = CURRENT_TIMESTAMP
            """,
            (so_o_clean, ma_npl_clean, int(is_active)),
        )
        conn.commit()
        conn.close()

    def _selected_result_so_o_ma_npl(self) -> tuple[str, str] | None:
        sel = self.result_tree.selection()
        if not sel:
            messagebox.showinfo("Thong bao", "Hay chon 1 dong NPL trong bang ket qua.")
            return None
        vals = self.result_tree.item(sel[0], "values")
        if not vals or len(vals) < 2:
            return None
        return str(vals[0]).strip(), str(vals[1]).strip()

    def add_utilization_from_selected(self) -> None:
        picked = self._selected_result_so_o_ma_npl()
        if not picked:
            return
        so_o, ma_npl = picked
        self._upsert_utilization_rule(so_o, ma_npl, 1)
        self.refresh_utilization_rules()
        self.reapply_current_views()
        messagebox.showinfo("Thanh cong", f"Da them Tan dung: {so_o} - {ma_npl}")

    def deactivate_utilization_from_selected(self) -> None:
        picked = self._selected_result_so_o_ma_npl()
        if not picked:
            return
        so_o, ma_npl = picked
        self._upsert_utilization_rule(so_o, ma_npl, 0)
        self.refresh_utilization_rules()
        self.reapply_current_views()
        messagebox.showinfo("Thanh cong", f"Da UnActive Tan dung: {so_o} - {ma_npl}")

    def add_utilization_manual(self) -> None:
        so_o = simpledialog.askstring("So O", "Nhap So O can Tan dung:")
        if so_o is None:
            return
        ma_npl = simpledialog.askstring("Ma NPL", "Nhap Ma NPL can Tan dung:")
        if ma_npl is None:
            return
        self._upsert_utilization_rule(so_o, ma_npl, 1)
        self.refresh_utilization_rules()
        self.reapply_current_views()

    def _selected_rule_pair(self) -> tuple[str, str] | None:
        sel = self.rules_tree.selection()
        if not sel:
            messagebox.showinfo("Thong bao", "Chon 1 dong quy tac.")
            return None
        vals = self.rules_tree.item(sel[0], "values")
        if not vals or len(vals) < 2:
            return None
        return str(vals[0]).strip(), str(vals[1]).strip()

    def activate_selected_rule(self) -> None:
        pair = self._selected_rule_pair()
        if not pair:
            return
        self._upsert_utilization_rule(pair[0], pair[1], 1)
        self.refresh_utilization_rules()
        self.reapply_current_views()

    def deactivate_selected_rule(self) -> None:
        pair = self._selected_rule_pair()
        if not pair:
            return
        self._upsert_utilization_rule(pair[0], pair[1], 0)
        self.refresh_utilization_rules()
        self.reapply_current_views()

    def refresh_utilization_rules(self) -> None:
        if not hasattr(self, "rules_tree"):
            return
        for item in self.rules_tree.get_children():
            self.rules_tree.delete(item)
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        rows = cur.execute(
            """
            SELECT so_o, ma_npl, is_active, COALESCE(created_at, ''), COALESCE(updated_at, '')
            FROM npl_utilization_rules
            ORDER BY updated_at DESC, id DESC
            """
        ).fetchall()
        conn.close()
        for so_o, ma_npl, active, created_at, updated_at in rows:
            self.rules_tree.insert(
                "",
                "end",
                values=(so_o, ma_npl, "Co" if int(active) == 1 else "Khong", created_at, updated_at),
            )

    def reapply_current_views(self) -> None:
        if self.last_result_base_df is not None and not self.last_result_base_df.empty:
            self.last_result_df = self._apply_utilization_rules(self.last_result_base_df)
            self.render_result(self.result_tree, self.last_result_df)
            if self.loaded_run_id is not None:
                self.render_result(self.db_item_tree, self.last_result_df)
        if self.last_summary_source_df is not None and not self.last_summary_source_df.empty:
            self._render_summary_from_source_df(self.last_summary_source_df, self.last_summary_run_ids)

    def export_excel(self) -> None:
        if self.last_result_df is None or self.last_result_df.empty:
            messagebox.showwarning("Chua co du lieu", "Hay bam 'Kiem tra' truoc khi xuat.")
            return
        output_path = filedialog.asksaveasfilename(
            title="Luu file ket qua",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"ket_qua_kiem_tra_{self.o_number_var.get().strip() or 'so_o'}.xlsx",
        )
        if not output_path:
            return
        export_df = self.last_result_df.copy()
        self._write_excel_with_ket_luan_color(export_df, output_path, "KetQua")
        trace_df = pd.DataFrame(
            [
                {"field": "bom_file", "value": self.bom_file_var.get().strip()},
                {"field": "stock_file", "value": self.stock_file_var.get().strip()},
                {"field": "trace_bom_a6", "value": self._current_trace_bom_a6()},
                {"field": "trace_stock_a6", "value": self._current_trace_stock_a6()},
            ]
        )
        self._write_trace_sheet(output_path, trace_df, "Trace")
        self.status_label.config(text=f"Da xuat file: {output_path}")
        messagebox.showinfo("Thanh cong", f"Da xuat file Excel:\n{output_path}")

    def save_current_run(self) -> None:
        if self.last_result_df is None or self.last_result_df.empty:
            messagebox.showwarning("Chua co du lieu", "Khong co du lieu de save.")
            return
        run_name = simpledialog.askstring("Ten run", "Nhap ten run de luu:", initialvalue=f"Run {self.o_number_var.get().strip()}")
        if run_name is None:
            return
        note = simpledialog.askstring("Ghi chu", "Nhap ghi chu (co the bo trong):", initialvalue="")
        trace_bom_a6 = self._current_trace_bom_a6()
        trace_stock_a6 = self._current_trace_stock_a6()
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO runs (run_name, so_o, bom_file, stock_file, note, trace_bom_a6, trace_stock_a6)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_name.strip() or f"Run {self.o_number_var.get().strip()}",
                self.o_number_var.get().strip(),
                self.bom_file_var.get().strip(),
                self.stock_file_var.get().strip(),
                (note or "").strip(),
                trace_bom_a6,
                trace_stock_a6,
            ),
        )
        run_id = cur.lastrowid
        rows = [
            (
                run_id,
                str(r["so_o"]),
                str(r["ma_npl"]),
                str(r["ten_npl"]),
                str(r.get("hang_nhap", "")).strip(),
                None if pd.isna(r["ton_thuc_te"]) else float(r["ton_thuc_te"]),
                None if pd.isna(r["ton_dm_chua_xuat"]) else float(r["ton_dm_chua_xuat"]),
                None if pd.isna(r["so_luong_can"]) else float(r["so_luong_can"]),
                str(r["ket_luan"]),
            )
            for _, r in self.last_result_df.iterrows()
        ]
        cur.executemany(
            """
            INSERT INTO run_items
            (run_id, so_o, ma_npl, ten_npl, hang_nhap, ton_thuc_te, ton_dm_chua_xuat, so_luong_can, ket_luan)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        conn.commit()
        conn.close()
        self.refresh_runs()
        messagebox.showinfo("Thanh cong", f"Da luu database, run_id = {run_id}")

    def refresh_runs(self) -> None:
        for item in self.runs_tree.get_children():
            self.runs_tree.delete(item)
        kw = normalize_text(self.search_db_var.get().strip())
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT r.id, r.run_name, r.so_o, r.created_at, r.note, COUNT(i.id) AS item_count
            FROM runs r
            LEFT JOIN run_items i ON i.run_id = r.id
            GROUP BY r.id, r.run_name, r.so_o, r.created_at, r.note
            ORDER BY r.id DESC
            """
        )
        rows = cur.fetchall()
        conn.close()
        for row in rows:
            row_text = " ".join([normalize_text(v) for v in row if v is not None])
            if kw and kw not in row_text:
                continue
            self.runs_tree.insert("", "end", values=row)
        self.refresh_summary_run_list()

    def refresh_summary_run_list(self) -> None:
        if not hasattr(self, "summary_runs_listbox"):
            return
        self.summary_runs_listbox.delete(0, tk.END)
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, run_name, so_o, created_at
            FROM runs
            ORDER BY id DESC
            """
        )
        rows = cur.fetchall()
        conn.close()
        for run_id, run_name, so_o, created_at in rows:
            display_text = f"{run_id} | {run_name} | {so_o} | {created_at}"
            self.summary_runs_listbox.insert(tk.END, display_text)

    def select_all_summary_runs(self) -> None:
        if not hasattr(self, "summary_runs_listbox"):
            return
        self.summary_runs_listbox.select_set(0, tk.END)

    def clear_summary_run_selection(self) -> None:
        if not hasattr(self, "summary_runs_listbox"):
            return
        self.summary_runs_listbox.selection_clear(0, tk.END)

    def _selected_summary_run_ids(self) -> list[int]:
        selected_indices = self.summary_runs_listbox.curselection()
        run_ids: list[int] = []
        for idx in selected_indices:
            text = self.summary_runs_listbox.get(idx)
            run_id_text = text.split("|", 1)[0].strip()
            if run_id_text.isdigit():
                run_ids.append(int(run_id_text))
        return run_ids

    def _selected_run_id(self) -> int | None:
        selected = self.runs_tree.selection()
        if not selected:
            messagebox.showwarning("Chua chon run", "Hay chon 1 run trong bang.")
            return None
        return int(self.runs_tree.item(selected[0], "values")[0])

    def load_selected_run(self) -> None:
        run_id = self._selected_run_id()
        if run_id is None:
            return
        conn = sqlite3.connect(DB_FILE)
        query = """
            SELECT so_o, ma_npl, ten_npl, COALESCE(hang_nhap, '') AS hang_nhap,
                   ton_thuc_te, ton_dm_chua_xuat, COALESCE(so_luong_can, 0) AS so_luong_can, ket_luan
            FROM run_items
            WHERE run_id = ?
            ORDER BY id
        """
        df = pd.read_sql_query(query, conn, params=(run_id,))
        cur = conn.cursor()
        cur.execute("SELECT so_o, bom_file, stock_file FROM runs WHERE id = ?", (run_id,))
        run_meta = cur.fetchone()
        cur.execute("SELECT COALESCE(trace_bom_a6, ''), COALESCE(trace_stock_a6, '') FROM runs WHERE id = ?", (run_id,))
        trace_meta = cur.fetchone()
        conn.close()
        if df.empty:
            messagebox.showwarning("Trong", "Run nay khong co item.")
            return
        self.last_result_base_df = df.copy()
        self.last_result_df = self._apply_utilization_rules(df)
        self.loaded_run_id = run_id
        self.o_number_var.set(str(run_meta[0] if run_meta else ""))
        self.bom_file_var.set(str(run_meta[1] if run_meta else ""))
        self.stock_file_var.set(str(run_meta[2] if run_meta else ""))
        self.last_trace_bom_a6 = str(trace_meta[0] if trace_meta else "").strip()
        self.last_trace_stock_a6 = str(trace_meta[1] if trace_meta else "").strip()
        self.render_result(self.result_tree, self.last_result_df)
        self.render_result(self.db_item_tree, self.last_result_df)
        self.notebook.select(self.tab_check)
        self.status_label.config(text=f"Da load run_id={run_id}, so dong={len(df)}")

    def export_selected_run(self) -> None:
        run_id = self._selected_run_id()
        if run_id is None:
            return
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        run_row = cur.execute("SELECT * FROM runs WHERE id = ?", (run_id,)).fetchone()
        if run_row is None:
            conn.close()
            messagebox.showerror("Export run", "Khong tim thay run.")
            return
        item_rows = [
            dict(r)
            for r in cur.execute("SELECT * FROM run_items WHERE run_id = ? ORDER BY id", (run_id,)).fetchall()
        ]
        conn.close()

        payload = {
            "format": "npl_checker_run_export_v1",
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "run": dict(run_row),
            "run_items": item_rows,
        }
        out = filedialog.asksaveasfilename(
            title="Luu file run",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"npl_run_{run_id}.json",
        )
        if not out:
            return
        Path(out).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        messagebox.showinfo("Export run", f"Da xuat run_id={run_id}:\n{out}")

    def import_run_from_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Chon file run da export",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception as exc:
            messagebox.showerror("Import run", f"Khong doc duoc file JSON:\n{exc}")
            return
        if str(payload.get("format", "")).strip() != "npl_checker_run_export_v1":
            messagebox.showerror("Import run", "Sai dinh dang file export run NPL Checker.")
            return
        run = payload.get("run") or {}
        items = payload.get("run_items") or []
        if not isinstance(run, dict) or not isinstance(items, list):
            messagebox.showerror("Import run", "Noi dung file khong hop le.")
            return

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO runs (run_name, so_o, bom_file, stock_file, note, created_at, trace_bom_a6, trace_stock_a6)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(run.get("run_name", "")).strip() or "Imported run",
                str(run.get("so_o", "")).strip(),
                str(run.get("bom_file", "")).strip(),
                str(run.get("stock_file", "")).strip(),
                str(run.get("note", "")).strip(),
                str(run.get("created_at", "")).strip() or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(run.get("trace_bom_a6", "")).strip(),
                str(run.get("trace_stock_a6", "")).strip(),
            ),
        )
        new_run_id = int(cur.lastrowid)

        item_rows: list[tuple] = []
        for r in items:
            if not isinstance(r, dict):
                continue
            ton_thuc_te = r.get("ton_thuc_te")
            ton_dm = r.get("ton_dm_chua_xuat")
            so_luong_can = r.get("so_luong_can")
            item_rows.append(
                (
                    new_run_id,
                    str(r.get("so_o", "")).strip(),
                    str(r.get("ma_npl", "")).strip(),
                    str(r.get("ten_npl", "")).strip(),
                    str(r.get("hang_nhap", "")).strip(),
                    None if ton_thuc_te in (None, "") else float(ton_thuc_te),
                    None if ton_dm in (None, "") else float(ton_dm),
                    None if so_luong_can in (None, "") else float(so_luong_can),
                    str(r.get("ket_luan", "")).strip(),
                )
            )
        if item_rows:
            cur.executemany(
                """
                INSERT INTO run_items
                (run_id, so_o, ma_npl, ten_npl, hang_nhap, ton_thuc_te, ton_dm_chua_xuat, so_luong_can, ket_luan)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                item_rows,
            )
        conn.commit()
        conn.close()
        self.refresh_runs()
        messagebox.showinfo("Import run", f"Da import thanh cong run moi: run_id={new_run_id}")

    def update_selected_run_note(self) -> None:
        run_id = self._selected_run_id()
        if run_id is None:
            return
        new_note = simpledialog.askstring("Update note", "Nhap ghi chu moi:")
        if new_note is None:
            return
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("UPDATE runs SET note = ? WHERE id = ?", (new_note.strip(), run_id))
        conn.commit()
        conn.close()
        self.refresh_runs()

    def delete_selected_run(self) -> None:
        run_id = self._selected_run_id()
        if run_id is None:
            return
        if not messagebox.askyesno("Xac nhan", f"Xoa run_id={run_id}?"):
            return
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("DELETE FROM run_items WHERE run_id = ?", (run_id,))
        cur.execute("DELETE FROM runs WHERE id = ?", (run_id,))
        conn.commit()
        conn.close()
        self.refresh_runs()
        for item in self.db_item_tree.get_children():
            self.db_item_tree.delete(item)

    def build_summary_report(self) -> None:
        selected_run_ids = self._selected_summary_run_ids()
        if not selected_run_ids:
            messagebox.showwarning(
                "Chua chon run",
                "Hay chon it nhat 1 run trong danh sach de tao Super Report.",
            )
            return

        conn = sqlite3.connect(DB_FILE)
        placeholders = ",".join("?" for _ in selected_run_ids)
        query = """
            SELECT
                ma_npl,
                ten_npl,
                COALESCE(hang_nhap, '') AS hang_nhap,
                so_o,
                run_id,
                ton_thuc_te,
                ton_dm_chua_xuat,
                COALESCE(so_luong_can, 0) AS so_luong_can
            FROM run_items
            WHERE run_id IN ({run_ids})
            ORDER BY ma_npl, so_o, run_id, id
        """.format(run_ids=placeholders)
        df = pd.read_sql_query(query, conn, params=selected_run_ids)
        conn.close()
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        if df.empty:
            self.last_summary_df = None
            self.last_summary_source_df = None
            self.summary_label.config(text="Chua co du lieu trong database.")
            return
        self.last_summary_source_df = df.copy()
        self._render_summary_from_source_df(df, selected_run_ids)

    def _render_summary_from_source_df(self, df_source: pd.DataFrame, run_ids: list[int]) -> None:
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        df = df_source.copy()
        df["ma_npl"] = df["ma_npl"].fillna("").astype(str)
        df["ten_npl"] = df["ten_npl"].fillna("").astype(str)
        df["hang_nhap"] = df["hang_nhap"].fillna("").astype(str)
        df["so_o"] = df["so_o"].fillna("").astype(str)
        df["ton_thuc_te"] = pd.to_numeric(df["ton_thuc_te"], errors="coerce")
        df["ton_dm_chua_xuat"] = pd.to_numeric(df["ton_dm_chua_xuat"], errors="coerce")
        df["so_luong_can"] = pd.to_numeric(df["so_luong_can"], errors="coerce").fillna(0.0)
        df["tan_dung"] = "Khong"

        keys = self._active_utilization_keys()
        if keys:
            pair = list(zip(normalize_ma_series(df["so_o"]), normalize_ma_series(df["ma_npl"])))
            util_mask = pd.Series([(so_o, ma) in keys for so_o, ma in pair], index=df.index)
            df.loc[util_mask, "so_luong_can"] = 0.0
            df.loc[util_mask, "tan_dung"] = "Co"

        grouped = (
            df.groupby(["ma_npl", "ten_npl"], as_index=False)
            .agg(
                hang_nhap=("hang_nhap", select_hang_nhap_label),
                ton_thuc_te=("ton_thuc_te", "sum"),
                ton_dm_chua_xuat=("ton_dm_chua_xuat", "sum"),
                so_luong_can=("so_luong_can", "sum"),
                tan_dung=("tan_dung", summarize_tan_dung),
                so_o=("so_o", lambda s: ", ".join(sorted({x for x in s if str(x).strip()}))),
                run_id=("run_id", lambda s: ", ".join(str(int(x)) for x in sorted({int(x) for x in s if pd.notna(x)}))),
            )
            .copy()
        )
        grouped["ket_luan"] = classify_stock_status_series(
            grouped["ton_thuc_te"],
            grouped["ton_dm_chua_xuat"],
            grouped["so_luong_can"],
        )

        grouped_npl = int(grouped["ma_npl"].nunique())
        neg_count = int((grouped["ket_luan"] == "Thieu NPL").sum())
        self.last_summary_df = grouped.copy()
        self.last_summary_run_ids = list(run_ids)
        self.summary_label.config(
            text=f"File gop ({len(run_ids)} run): {len(grouped)} dong tong hop, {grouped_npl} ma NPL, {neg_count} ma thieu NPL."
        )

        for _, row in grouped.iterrows():
            tag = ket_luan_tag(str(row["ket_luan"]))
            self.summary_tree.insert(
                "",
                "end",
                values=(
                    row["ma_npl"],
                    row["ten_npl"],
                    row["hang_nhap"],
                    row["so_o"],
                    row["run_id"],
                    "" if pd.isna(row["ton_thuc_te"]) else f"{row['ton_thuc_te']:.4f}",
                    ""
                    if pd.isna(row["ton_dm_chua_xuat"])
                    else f"{row['ton_dm_chua_xuat']:.4f}",
                    "" if pd.isna(row["so_luong_can"]) else f"{row['so_luong_can']:.4f}",
                    row["tan_dung"],
                    row["ket_luan"],
                ),
                tags=(tag,),
            )

    def export_summary_excel(self) -> None:
        if self.last_summary_df is None or self.last_summary_df.empty:
            messagebox.showwarning(
                "Chua co du lieu",
                "Hay bam 'Tao bao cao tu run da chon' truoc khi Export Excel.",
            )
            return
        output_path = filedialog.asksaveasfilename(
            title="Luu file Super Report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="super_report.xlsx",
        )
        if not output_path:
            return
        export_df = self.last_summary_df.copy()
        self._write_excel_with_ket_luan_color(export_df, output_path, "SuperReport")
        trace_rows = self._summary_trace_rows(self.last_summary_run_ids)
        if trace_rows:
            self._write_trace_sheet(output_path, pd.DataFrame(trace_rows), "TraceRuns")
        messagebox.showinfo("Thanh cong", f"Da xuat file Super Report:\n{output_path}")

    def _write_excel_with_ket_luan_color(
        self,
        export_df: pd.DataFrame,
        output_path: str,
        sheet_name: str,
    ) -> None:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            red_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
            orange_fill = PatternFill(start_color="FFE6BF", end_color="FFE6BF", fill_type="solid")
            green_fill = PatternFill(start_color="E7F7E7", end_color="E7F7E7", fill_type="solid")
            if "ket_luan" not in export_df.columns:
                return
            ket_luan_col_idx = list(export_df.columns).index("ket_luan") + 1
            for row_idx in range(2, len(export_df) + 2):
                status = str(ws.cell(row=row_idx, column=ket_luan_col_idx).value or "").strip()
                if status == "Thieu NPL":
                    fill = red_fill
                elif status == "Chu y":
                    fill = orange_fill
                else:
                    fill = green_fill
                for col_idx in range(1, len(export_df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    def _current_trace_bom_a6(self) -> str:
        if self.last_trace_bom_a6:
            return self.last_trace_bom_a6
        self.last_trace_bom_a6 = trace_sheet_a6(self.bom_file_var.get().strip())
        return self.last_trace_bom_a6

    def _current_trace_stock_a6(self) -> str:
        if self.last_trace_stock_a6:
            return self.last_trace_stock_a6
        self.last_trace_stock_a6 = trace_sheet_a6(self.stock_file_var.get().strip())
        return self.last_trace_stock_a6

    def _write_trace_sheet(self, output_path: str, trace_df: pd.DataFrame, sheet_name: str) -> None:
        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            trace_df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _summary_trace_rows(self, run_ids: list[int]) -> list[dict]:
        if not run_ids:
            return []
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        out: list[dict] = []
        for rid in run_ids:
            row = cur.execute(
                """
                SELECT COALESCE(run_name, ''), COALESCE(so_o, ''), COALESCE(created_at, ''),
                       COALESCE(trace_bom_a6, ''), COALESCE(trace_stock_a6, '')
                FROM runs
                WHERE id = ?
                """,
                (rid,),
            ).fetchone()
            if not row:
                continue
            out.append(
                {
                    "run_id": rid,
                    "run_name": row[0],
                    "so_o": row[1],
                    "created_at": row[2],
                    "trace_bom_a6": row[3],
                    "trace_stock_a6": row[4],
                }
            )
        conn.close()
        return out

    def switch_to_summary(self) -> None:
        self.build_summary_report()
        self.notebook.select(self.tab_summary)


def main(back_to_launcher: callable | None = None) -> None:
    root = tk.Tk()
    app = App(root, back_to_launcher=back_to_launcher)
    root.mainloop()


if __name__ == "__main__":
    main()
